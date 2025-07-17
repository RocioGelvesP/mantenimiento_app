from flask import Blueprint, render_template, request, url_for, flash, redirect, send_file, make_response, jsonify
from flask_login import login_required, current_user
from flask import current_app
from datetime import datetime, timezone, timedelta
from models import db, Equipo, MotorEquipo, Programado, Company, EquipoMedicion, HistorialEquipo, get_or_404
from flask_wtf import FlaskForm
from wtforms import StringField, BooleanField
from utils import require_role, require_any_role, get_equipos_filtrados_por_rol, require_delete_permission, get_pdf_config, get_pdf_options, registrar_auditoria, generar_y_enviar_pdf_ficha_tecnica
import pandas as pd
import pdfkit
import shutil
import io, os
import traceback
from forms import EquipoForm
from werkzeug.utils import secure_filename

bp = Blueprint('equipment', __name__, url_prefix='/equipos')

def get_local_datetime():
    """Obtiene la fecha y hora local (Colombia)"""
    utc_now = datetime.now(timezone.utc)
    colombia_tz = timezone(timedelta(hours=-5))  # UTC-5 para Colombia
    return utc_now.astimezone(colombia_tz)

@bp.route('/listar', methods=['GET'])
@login_required
def listar_equipos():
    # Obtener parÃ¡metros de filtro desde la URL
    codigo = request.args.get('codigo', '')
    nombre = request.args.get('nombre', '')
    fecha_ingreso = request.args.get('fecha_ingreso', '')
    ubicacion = request.args.get('ubicacion', '')
    estado = request.args.get('estado_eq', '')

    # Obtener consulta base filtrada por rol
    query = get_equipos_filtrados_por_rol()

    if codigo:
        query = query.filter(Equipo.codigo.ilike(f"%{codigo}%"))
    if nombre:
        query = query.filter(Equipo.nombre.ilike(f"%{nombre}%"))
    if fecha_ingreso:
        query = query.filter(Equipo.fecha_ingreso == fecha_ingreso)
    if ubicacion:
        query = query.filter(Equipo.ubicacion.ilike(f"%{ubicacion}%"))
    if estado:
        query = query.filter(Equipo.estado_eq.ilike(f"%{estado}%"))

    equipos = query.all()
    
    # Calcular informaciÃ³n de mantenimientos para cada equipo
    for equipo in equipos:
        # Buscar el Ãºltimo mantenimiento realizado para este equipo
        ultimo_mantenimiento = Programado.query.filter(
            Programado.codigo == equipo.codigo,
            Programado.estado_final == 'Completado'
        ).order_by(Programado.fecha_prog.desc()).first()
        
        # Asignar informaciÃ³n del Ãºltimo mantenimiento
        if ultimo_mantenimiento:
            equipo.ultimo_tipo_mant = ultimo_mantenimiento.tipo_mantenimiento
            equipo.ultima_fecha_mant = ultimo_mantenimiento.fecha_prog
        else:
            equipo.ultimo_tipo_mant = None
            equipo.ultima_fecha_mant = None
        
        # Buscar el prÃ³ximo mantenimiento programado para este equipo
        proximo_mantenimiento = Programado.query.filter(
            Programado.codigo == equipo.codigo,
            Programado.fecha_prog >= datetime.now().date(),
            Programado.estado_inicial.in_(['Programado', 'Asignado'])
        ).order_by(Programado.fecha_prog.asc()).first()
        
        # Asignar informaciÃ³n del prÃ³ximo mantenimiento
        if proximo_mantenimiento:
            equipo.proximo_tipo_mant = proximo_mantenimiento.tipo_mantenimiento
            equipo.fecha_prox_mant = proximo_mantenimiento.fecha_prog
        else:
            equipo.proximo_tipo_mant = None
            equipo.fecha_prox_mant = None

    # Calcular contadores sobre la lista filtrada
    equipos_activos = [e for e in equipos if e.estado_eq and e.estado_eq.strip().lower() == 'activo']
    equipos_de_baja = [e for e in equipos if e.estado_eq and e.estado_eq.strip().lower() == 'de baja']
    equipos_inactivos = [e for e in equipos if not e.estado_eq or e.estado_eq.strip().lower() not in ['activo', 'de baja']]

    # Para usuarios con roles superiores, tambiÃ©n mostrar contadores de todos los equipos del sistema
    total_activos_sistema = None
    total_de_baja_sistema = None
    total_inactivos_sistema = None
    
    if current_user.is_super_admin() or current_user.is_admin() or current_user.is_supervisor():
        # Obtener todos los equipos del sistema para mostrar contadores completos
        todos_equipos_sistema = Equipo.query.all()
        total_activos_sistema = len([e for e in todos_equipos_sistema if e.estado_eq and e.estado_eq.strip().lower() == 'activo'])
        total_de_baja_sistema = len([e for e in todos_equipos_sistema if e.estado_eq and e.estado_eq.strip().lower() == 'de baja'])
        total_inactivos_sistema = len([e for e in todos_equipos_sistema if not e.estado_eq or e.estado_eq.strip().lower() not in ['activo', 'de baja']])

    return render_template('equipos/listar_equipos.html', 
                         equipos=equipos, 
                         equipos_activos=equipos_activos, 
                         equipos_inactivos=equipos_inactivos, 
                         equipos_de_baja=equipos_de_baja,
                         total_activos_sistema=total_activos_sistema,
                         total_de_baja_sistema=total_de_baja_sistema,
                         total_inactivos_sistema=total_inactivos_sistema)

    #NUEVO EQUIPO
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
@bp.route('/nuevo', methods=['GET', 'POST'])
@login_required
@require_any_role('super_admin', 'admin', 'supervisor')
def nuevo_equipo():
    form = EquipoForm()
    if form.validate_on_submit():
        try:
            # Crear equipo
            equipo = Equipo(
                codigo=form.codigo.data,
                nombre=form.nombre.data,
                fecha_ingreso=form.fecha_ingreso.data,
                registro_nuevo=form.registro_nuevo.data,
                actualizacion=form.actualizacion.data,
                num_fabricacion=form.num_fabricacion.data,
                fabricante=form.fabricante.data,
                nom_contacto=form.nom_contacto.data,
                telefono=form.telefono.data,
                propietario=form.propietario.data,
                tipo_eq=form.tipo_eq.data,
                modelo=form.modelo.data,
                tipo_control=form.tipo_control.data,
                ubicacion=form.ubicacion.data,
                clase=form.clase.data,
                marca=form.marca.data,
                referencia=form.referencia.data,
                serie=form.serie.data,
                color=form.color.data,
                altura=form.altura.data,
                largo=form.largo.data,
                ancho=form.ancho.data,
                peso=form.peso.data,
                corriente=form.corriente.data,
                potencia=form.potencia.data,
                voltaje=form.voltaje.data,
                tipo_refrig=form.tipo_refrig.data,
                tipo_comb=form.tipo_comb.data,
                tipo_lub=form.tipo_lub.data,
                repuestos=form.repuestos.data,
                estado_eq=form.estado_eq.data,
                n_motores=form.n_motores.data,
                observaciones=form.observaciones.data,
                hist_mtto=form.hist_mtto.data,
                funcion_maq=form.funcion_maq.data,
                acciones=form.acciones.data,
                company_id=form.company_id.data,
                instructivos_file=None,
                estandar_seguridad_file=None,
                operacion_file=None,
                mecanico_file=None,
                electrico_file=None,
                partes_file=None,
                proceso=form.proceso.data,
                anios_operacion=form.anios_operacion.data,
                fecha_fabricacion=form.fecha_fabricacion.data,
                tipo_instalacion=form.tipo_instalacion.data,
                cartas_lubricacion=form.cartas_lubricacion.data,
                metodo_codificacion=form.metodo_codificacion.data,
                frecuencia_mantenimiento=form.frecuencia_mantenimiento.data,
                ficha_tecnica=form.ficha_tecnica.data,
                hoja_vida=form.hoja_vida.data,
                es_equipo=form.es_equipo.data,
                es_maquina=form.es_maquina.data,
                preoperacional=form.preoperacional.data,
                plan_mantenimiento=form.plan_mantenimiento.data,
                inspeccion_seguridad=form.inspeccion_seguridad.data,
                procedimientos_operacion=form.procedimientos_operacion.data,
                manual_usuario=form.manual_usuario.data,
                certificaciones=form.certificaciones.data,
                registro_mantenimientos=form.registro_mantenimientos.data
            )
            equipo.set_tipo_energia(form.tipo_energia.data)
            db.session.add(equipo)
            db.session.flush()

            # Manejar archivos
            if 'imagen' in request.files and request.files['imagen'].filename:
                file = request.files['imagen']
                if file and allowed_file(file.filename):
                    filename = secure_filename(f"{equipo.codigo}_{file.filename}")
                    file_path = os.path.join('static', 'uploads', filename)
                    file.save(file_path)
                    equipo.imagen = f'uploads/{filename}'
            
            if 'imagen_lubricacion' in request.files and request.files['imagen_lubricacion'].filename:
                file = request.files['imagen_lubricacion']
                if file and allowed_file(file.filename):
                    filename = secure_filename(f"{equipo.codigo}_lubricacion_{file.filename}")
                    file_path = os.path.join('static', 'uploads', filename)
                    file.save(file_path)
                    equipo.imagen_lubricacion = f'uploads/{filename}'
            
            # Manejar otros archivos
            archivos = ['instructivos_file', 'estandar_seguridad_file', 'operacion_file', 
                       'mecanico_file', 'electrico_file', 'partes_file']
            for archivo in archivos:
                if archivo in request.files and request.files[archivo].filename:
                    file = request.files[archivo]
                    if file and allowed_file(file.filename):
                        filename = secure_filename(f"{equipo.codigo}_{archivo}_{file.filename}")
                        file_path = os.path.join('static', 'uploads', 'documentos', filename)
                        os.makedirs(os.path.dirname(file_path), exist_ok=True)
                        file.save(file_path)
                        setattr(equipo, archivo, f'uploads/documentos/{filename}')

            # Historial de creaciÃ³n de campos de equipo
            for field in equipo.__table__.columns.keys():
                if field == 'codigo':
                    continue
                valor = getattr(equipo, field)
                if valor is not None and valor != '':
                    db.session.add(HistorialEquipo(
                        equipo_codigo=equipo.codigo,
                        tipo_cambio='creacion',
                        campo_modificado=field,
                        valor_anterior=None,
                        valor_nuevo=str(valor),
                        usuario=current_user.username if current_user.is_authenticated else None,
                        observaciones=f'CreaciÃ³n inicial del campo {field}',
                        fecha_cambio=get_local_datetime()
                    ))

            # Guardar motores y registrar historial de cada campo
            id = request.form.getlist('id[]')
            nomb_Motor = request.form.getlist('nomb_Motor[]')
            descrip_Motor = request.form.getlist('descrip_Motor[]')
            tipo_Motor = request.form.getlist('tipo_Motor[]')
            rotacion = request.form.getlist('rotacion[]')
            corriente_Motor = request.form.getlist('corriente_Motor[]')
            potencia_Motor = request.form.getlist('potencia_Motor[]')
            voltaje_Motor = request.form.getlist('voltaje_Motor[]')
            rpm_Motor = request.form.getlist('rpm_Motor[]')
            eficiencia = request.form.getlist('eficiencia[]')
            
            # Debug: Imprimir los datos recibidos
            print(f"DEBUG - Motores recibidos:")
            print(f"nomb_Motor: {nomb_Motor}")
            print(f"descrip_Motor: {descrip_Motor}")
            print(f"tipo_Motor: {tipo_Motor}")
            print(f"rotacion: {rotacion}")
            print(f"corriente_Motor: {corriente_Motor}")
            print(f"potencia_Motor: {potencia_Motor}")
            print(f"voltaje_Motor: {voltaje_Motor}")
            print(f"rpm_Motor: {rpm_Motor}")
            print(f"eficiencia: {eficiencia}")
            
            max_len = max(len(nomb_Motor), len(descrip_Motor), len(tipo_Motor), len(potencia_Motor), len(voltaje_Motor), len(corriente_Motor), len(rpm_Motor), len(eficiencia), len(rotacion))
            nomb_Motor.extend([''] * (max_len - len(nomb_Motor)))
            descrip_Motor.extend([''] * (max_len - len(descrip_Motor)))
            tipo_Motor.extend([''] * (max_len - len(tipo_Motor)))
            potencia_Motor.extend([''] * (max_len - len(potencia_Motor)))
            voltaje_Motor.extend([''] * (max_len - len(voltaje_Motor)))
            corriente_Motor.extend([''] * (max_len - len(corriente_Motor)))
            rpm_Motor.extend([''] * (max_len - len(rpm_Motor)))
            eficiencia.extend([''] * (max_len - len(eficiencia)))
            rotacion.extend([''] * (max_len - len(rotacion)))
            for i in range(len(nomb_Motor)):
                if nomb_Motor[i].strip():
                    motor = MotorEquipo(
                        equipo_codigo=equipo.codigo,
                        nomb_Motor=nomb_Motor[i],
                        descrip_Motor=descrip_Motor[i],
                        tipo_Motor=tipo_Motor[i],
                        potencia_Motor=potencia_Motor[i],
                        voltaje_Motor=voltaje_Motor[i],
                        corriente_Motor=corriente_Motor[i],
                        rpm_Motor=rpm_Motor[i],
                        eficiencia=eficiencia[i],
                        rotacion=rotacion[i]
                    )
                    db.session.add(motor)
                    db.session.flush()
                    for field in ['nomb_Motor','descrip_Motor','tipo_Motor','potencia_Motor','voltaje_Motor','corriente_Motor','rpm_Motor','eficiencia','rotacion']:
                        valor = getattr(motor, field)
                        if valor is not None and valor != '':
                            db.session.add(HistorialEquipo(
                                equipo_codigo=equipo.codigo,
                                tipo_cambio='creacion',
                                campo_modificado=f'motor_{field}',
                                valor_anterior=None,
                                valor_nuevo=str(valor),
                                usuario=current_user.username if current_user.is_authenticated else None,
                                observaciones=f'CreaciÃ³n inicial del campo {field} del motor',
                                fecha_cambio=get_local_datetime()
                            ))
            # Guardar equipos de mediciÃ³n y registrar historial de cada campo
            codigo_medicion = request.form.getlist('codigo_medicion[]')
            nombre_medicion = request.form.getlist('nombre_medicion[]')
            ubicacion_medicion = request.form.getlist('ubicacion_medicion[]')
            max_len_med = max(len(codigo_medicion), len(nombre_medicion), len(ubicacion_medicion))
            codigo_medicion.extend([''] * (max_len_med - len(codigo_medicion)))
            nombre_medicion.extend([''] * (max_len_med - len(nombre_medicion)))
            ubicacion_medicion.extend([''] * (max_len_med - len(ubicacion_medicion)))
            for i in range(max_len_med):
                if codigo_medicion[i].strip() or nombre_medicion[i].strip():
                    medicion = EquipoMedicion(
                        equipo_codigo=equipo.codigo,
                        codigo=codigo_medicion[i],
                        nombre=nombre_medicion[i],
                        ubicacion=ubicacion_medicion[i]
                    )
                    db.session.add(medicion)
                    db.session.flush()
                    for field in ['codigo','nombre','ubicacion']:
                        valor = getattr(medicion, field)
                        if valor is not None and valor != '':
                            db.session.add(HistorialEquipo(
                                equipo_codigo=equipo.codigo,
                                tipo_cambio='creacion',
                                campo_modificado=f'equipomedicion_{field}',
                                valor_anterior=None,
                                valor_nuevo=str(valor),
                                usuario=current_user.username if current_user.is_authenticated else None,
                                observaciones=f'CreaciÃ³n inicial del campo {field} del equipo de mediciÃ³n',
                                fecha_cambio=get_local_datetime()
                            ))
            db.session.commit()
            
            # Registrar auditoría
            registrar_auditoria(
                modulo='equipos',
                accion='CREAR',
                tabla='equipos',
                descripcion=f"Creó equipo {equipo.codigo} - {equipo.nombre}"
            )
            
            flash(f'MÃ¡quina/Equipo "{form.codigo.data}" y motores guardados correctamente.', 'success')
            return redirect(url_for('equipment.listar_equipos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el equipo: {str(e)}', 'danger')
            return redirect(url_for('equipment.nuevo_equipo'))
    if not request.form.get('n_motores'):
        form.n_motores.data = 0
    return render_template('equipos/nuevo_equipo.html', form=form)

    #EDITAR EQUIPO
@bp.route('/editar/<codigo>', methods=['GET', 'POST'])
@login_required
@require_any_role('super_admin', 'admin', 'supervisor')
def editar_equipo(codigo):
    equipo = get_or_404(Equipo, codigo)
    motores = MotorEquipo.query.filter_by(equipo_codigo=codigo).all()
    mediciones = EquipoMedicion.query.filter_by(equipo_codigo=codigo).all()
    
    # Debug: Imprimir los motores cargados
    print(f"DEBUG - Cargando motores para equipo {codigo}")
    print(f"DEBUG - Motores encontrados: {len(motores)}")
    for i, motor in enumerate(motores):
        print(f"DEBUG - Motor {i+1}: id={motor.id}, nomb_Motor='{motor.nomb_Motor}', descrip_Motor='{motor.descrip_Motor}'")
    
    form = EquipoForm(obj=equipo)
    
    if request.method == 'GET':
        form.tipo_energia.data = equipo.get_tipo_energia()
        # Forzar booleanos a False si vienen como None para que los checklists se muestren correctamente
        campos_booleanos = ['ficha_tecnica', 'hoja_vida', 'preoperacional', 'plan_mantenimiento', 'manuales', 'es_equipo', 'es_maquina', 'registro_nuevo', 'actualizacion']
        for campo in campos_booleanos:
            if hasattr(form, campo) and getattr(form, campo).data is None:
                getattr(form, campo).data = False
    
    if request.method == 'POST':
        try:
            form.n_motores.data = int(request.form.get('n_motores', 0))
        except (TypeError, ValueError):
            form.n_motores.data = 0

        if form.validate():
            try:
                # Guardar valores anteriores para comparaciÃ³n de equipo
                valores_anteriores = {col: getattr(equipo, col) for col in equipo.__table__.columns.keys() if col != 'codigo'}
                # Actualizar campos bÃ¡sicos
                form.populate_obj(equipo)
                equipo.set_tipo_energia(form.tipo_energia.data)
                
                # Manejar archivos
                if 'imagen' in request.files and request.files['imagen'].filename:
                    file = request.files['imagen']
                    if file and allowed_file(file.filename):
                        filename = secure_filename(f"{equipo.codigo}_{file.filename}")
                        file_path = os.path.join('static', 'uploads', filename)
                        file.save(file_path)
                        equipo.imagen = f'uploads/{filename}'
                
                if 'imagen_lubricacion' in request.files and request.files['imagen_lubricacion'].filename:
                    file = request.files['imagen_lubricacion']
                    if file and allowed_file(file.filename):
                        filename = secure_filename(f"{equipo.codigo}_lubricacion_{file.filename}")
                        file_path = os.path.join('static', 'uploads', filename)
                        file.save(file_path)
                        equipo.imagen_lubricacion = f'uploads/{filename}'
                
                # Manejar otros archivos
                archivos = ['instructivos_file', 'estandar_seguridad_file', 'operacion_file', 
                           'mecanico_file', 'electrico_file', 'partes_file']
                for archivo in archivos:
                    if archivo in request.files and request.files[archivo].filename:
                        file = request.files[archivo]
                        if file and allowed_file(file.filename):
                            filename = secure_filename(f"{equipo.codigo}_{archivo}_{file.filename}")
                            file_path = os.path.join('static', 'uploads', 'documentos', filename)
                            os.makedirs(os.path.dirname(file_path), exist_ok=True)
                            file.save(file_path)
                            setattr(equipo, archivo, f'uploads/documentos/{filename}')
                
                # Registrar cambios en el historial de equipo
                for campo, valor_anterior in valores_anteriores.items():
                    valor_nuevo = getattr(equipo, campo)
                    if str(valor_anterior) != str(valor_nuevo):
                        db.session.add(HistorialEquipo(
                            equipo_codigo=codigo,
                            tipo_cambio='actualizacion',
                            campo_modificado=campo,
                            valor_anterior=str(valor_anterior),
                            valor_nuevo=str(valor_nuevo),
                            usuario=current_user.username if current_user.is_authenticated else None,
                            observaciones=f'ActualizaciÃ³n del campo {campo}',
                            fecha_cambio=get_local_datetime()
                        ))
                # --- Motores ---
                id = request.form.getlist('id[]')
                nomb_Motor = request.form.getlist('nomb_Motor[]')
                descrip_Motor = request.form.getlist('descrip_Motor[]')
                tipo_Motor = request.form.getlist('tipo_Motor[]')
                rotacion = request.form.getlist('rotacion[]')
                corriente_Motor = request.form.getlist('corriente_Motor[]')
                potencia_Motor = request.form.getlist('potencia_Motor[]')
                voltaje_Motor = request.form.getlist('voltaje_Motor[]')
                rpm_Motor = request.form.getlist('rpm_Motor[]')
                eficiencia = request.form.getlist('eficiencia[]')
                
                # Debug: Imprimir los datos recibidos
                print(f"DEBUG - Motores recibidos:")
                print(f"nomb_Motor: {nomb_Motor}")
                print(f"descrip_Motor: {descrip_Motor}")
                print(f"tipo_Motor: {tipo_Motor}")
                print(f"rotacion: {rotacion}")
                print(f"corriente_Motor: {corriente_Motor}")
                print(f"potencia_Motor: {potencia_Motor}")
                print(f"voltaje_Motor: {voltaje_Motor}")
                print(f"rpm_Motor: {rpm_Motor}")
                print(f"eficiencia: {eficiencia}")
                
                max_len = max(len(nomb_Motor), len(descrip_Motor), len(tipo_Motor), len(potencia_Motor), len(voltaje_Motor), len(corriente_Motor), len(rpm_Motor), len(eficiencia), len(rotacion))
                nomb_Motor.extend([''] * (max_len - len(nomb_Motor)))
                descrip_Motor.extend([''] * (max_len - len(descrip_Motor)))
                tipo_Motor.extend([''] * (max_len - len(tipo_Motor)))
                potencia_Motor.extend([''] * (max_len - len(potencia_Motor)))
                voltaje_Motor.extend([''] * (max_len - len(voltaje_Motor)))
                corriente_Motor.extend([''] * (max_len - len(corriente_Motor)))
                rpm_Motor.extend([''] * (max_len - len(rpm_Motor)))
                eficiencia.extend([''] * (max_len - len(eficiencia)))
                rotacion.extend([''] * (max_len - len(rotacion)))
                # Comparar motores por Ã­ndice
                for i in range(max(max_len, len(motores))):
                    anterior = motores[i] if i < len(motores) else None
                    nuevo = {
                        'nomb_Motor': nomb_Motor[i] if i < len(nomb_Motor) else '',
                        'descrip_Motor': descrip_Motor[i] if i < len(descrip_Motor) else '',
                        'tipo_Motor': tipo_Motor[i] if i < len(tipo_Motor) else '',
                        'potencia_Motor': potencia_Motor[i] if i < len(potencia_Motor) else '',
                        'voltaje_Motor': voltaje_Motor[i] if i < len(voltaje_Motor) else '',
                        'corriente_Motor': corriente_Motor[i] if i < len(corriente_Motor) else '',
                        'rpm_Motor': rpm_Motor[i] if i < len(rpm_Motor) else '',
                        'eficiencia': eficiencia[i] if i < len(eficiencia) else '',
                        'rotacion': rotacion[i] if i < len(rotacion) else ''
                    }
                    if anterior:
                        for campo in nuevo.keys():
                            valor_ant = getattr(anterior, campo)
                            valor_nuevo = nuevo[campo]
                            if str(valor_ant) != str(valor_nuevo):
                                db.session.add(HistorialEquipo(
                                    equipo_codigo=codigo,
                                    tipo_cambio='actualizacion',
                                    campo_modificado=f'motor_{campo}_idx{i+1}',
                                    valor_anterior=str(valor_ant),
                                    valor_nuevo=str(valor_nuevo),
                                    usuario=current_user.username if current_user.is_authenticated else None,
                                    observaciones=f'ActualizaciÃ³n del campo {campo} del motor #{i+1}',
                                    fecha_cambio=get_local_datetime()
                                ))
                    else:
                        # Motor agregado
                        for campo, valor in nuevo.items():
                            if valor:
                                db.session.add(HistorialEquipo(
                                    equipo_codigo=codigo,
                                    tipo_cambio='creacion',
                                    campo_modificado=f'motor_{campo}_idx{i+1}',
                                    valor_anterior=None,
                                    valor_nuevo=str(valor),
                                    usuario=current_user.username if current_user.is_authenticated else None,
                                    observaciones=f'CreaciÃ³n del campo {campo} del motor #{i+1}',
                                    fecha_cambio=get_local_datetime()
                                ))
                # Motores eliminados
                if len(motores) > max_len:
                    for i in range(max_len, len(motores)):
                        anterior = motores[i]
                        for campo in ['nomb_Motor','descrip_Motor','tipo_Motor','potencia_Motor','voltaje_Motor','corriente_Motor','rpm_Motor','eficiencia','rotacion']:
                            valor_ant = getattr(anterior, campo)
                            if valor_ant:
                                db.session.add(HistorialEquipo(
                                    equipo_codigo=codigo,
                                    tipo_cambio='eliminacion',
                                    campo_modificado=f'motor_{campo}_idx{i+1}',
                                    valor_anterior=str(valor_ant),
                                    valor_nuevo=None,
                                    usuario=current_user.username if current_user.is_authenticated else None,
                                    observaciones=f'EliminaciÃ³n del campo {campo} del motor #{i+1}',
                                    fecha_cambio=get_local_datetime()
                                ))
                # Eliminar y agregar motores en la base
                MotorEquipo.query.filter_by(equipo_codigo=codigo).delete()
                print(f"DEBUG - Intentando guardar {max_len} motores")
                for i in range(max_len):
                    # Verificar si al menos un campo tiene datos
                    tiene_datos = any([
                        nomb_Motor[i].strip(),
                        descrip_Motor[i].strip(),
                        tipo_Motor[i].strip(),
                        potencia_Motor[i].strip(),
                        voltaje_Motor[i].strip(),
                        corriente_Motor[i].strip(),
                        rpm_Motor[i].strip(),
                        eficiencia[i].strip(),
                        rotacion[i].strip()
                    ])
                    
                    print(f"DEBUG - Motor {i+1}: tiene_datos={tiene_datos}, nomb_Motor='{nomb_Motor[i]}'")
                    
                    if tiene_datos:
                        motor = MotorEquipo(
                            equipo_codigo=codigo,
                            nomb_Motor=nomb_Motor[i],
                            descrip_Motor=descrip_Motor[i],
                            tipo_Motor=tipo_Motor[i],
                            potencia_Motor=potencia_Motor[i],
                            voltaje_Motor=voltaje_Motor[i],
                            corriente_Motor=corriente_Motor[i],
                            rpm_Motor=rpm_Motor[i],
                            eficiencia=eficiencia[i],
                            rotacion=rotacion[i]
                        )
                        db.session.add(motor)
                        print(f"DEBUG - Motor {i+1} agregado a la sesiÃ³n")
                    else:
                        print(f"DEBUG - Motor {i+1} no tiene datos, no se guarda")
                # --- Equipos de mediciÃ³n ---
                codigo_medicion = request.form.getlist('codigo_medicion[]')
                nombre_medicion = request.form.getlist('nombre_medicion[]')
                ubicacion_medicion = request.form.getlist('ubicacion_medicion[]')
                max_len_med = max(len(codigo_medicion), len(nombre_medicion), len(ubicacion_medicion))
                codigo_medicion.extend([''] * (max_len_med - len(codigo_medicion)))
                nombre_medicion.extend([''] * (max_len_med - len(nombre_medicion)))
                ubicacion_medicion.extend([''] * (max_len_med - len(ubicacion_medicion)))
                for i in range(max(max_len_med, len(mediciones))):
                    anterior = mediciones[i] if i < len(mediciones) else None
                    nuevo = {
                        'codigo': codigo_medicion[i] if i < len(codigo_medicion) else '',
                        'nombre': nombre_medicion[i] if i < len(nombre_medicion) else '',
                        'ubicacion': ubicacion_medicion[i] if i < len(ubicacion_medicion) else ''
                    }
                    if anterior:
                        for campo in nuevo.keys():
                            valor_ant = getattr(anterior, campo)
                            valor_nuevo = nuevo[campo]
                            if str(valor_ant) != str(valor_nuevo):
                                db.session.add(HistorialEquipo(
                                    equipo_codigo=codigo,
                                    tipo_cambio='actualizacion',
                                    campo_modificado=f'equipomedicion_{campo}_idx{i+1}',
                                    valor_anterior=str(valor_ant),
                                    valor_nuevo=str(valor_nuevo),
                                    usuario=current_user.username if current_user.is_authenticated else None,
                                    observaciones=f'ActualizaciÃ³n del campo {campo} del equipo de mediciÃ³n #{i+1}',
                                    fecha_cambio=get_local_datetime()
                                ))
                    else:
                        # Equipo de mediciÃ³n agregado
                        for campo, valor in nuevo.items():
                            if valor:
                                db.session.add(HistorialEquipo(
                                    equipo_codigo=codigo,
                                    tipo_cambio='creacion',
                                    campo_modificado=f'equipomedicion_{campo}_idx{i+1}',
                                    valor_anterior=None,
                                    valor_nuevo=str(valor),
                                    usuario=current_user.username if current_user.is_authenticated else None,
                                    observaciones=f'CreaciÃ³n del campo {campo} del equipo de mediciÃ³n #{i+1}',
                                    fecha_cambio=get_local_datetime()
                                ))
                # Equipos de mediciÃ³n eliminados
                if len(mediciones) > max_len_med:
                    for i in range(max_len_med, len(mediciones)):
                        anterior = mediciones[i]
                        for campo in ['codigo','nombre','ubicacion']:
                            valor_ant = getattr(anterior, campo)
                            if valor_ant:
                                db.session.add(HistorialEquipo(
                                    equipo_codigo=codigo,
                                    tipo_cambio='eliminacion',
                                    campo_modificado=f'equipomedicion_{campo}_idx{i+1}',
                                    valor_anterior=str(valor_ant),
                                    valor_nuevo=None,
                                    usuario=current_user.username if current_user.is_authenticated else None,
                                    observaciones=f'EliminaciÃ³n del campo {campo} del equipo de mediciÃ³n #{i+1}',
                                    fecha_cambio=get_local_datetime()
                                ))
                # Eliminar y agregar equipos de mediciÃ³n en la base
                EquipoMedicion.query.filter_by(equipo_codigo=codigo).delete()
                for i in range(max_len_med):
                    if codigo_medicion[i].strip() or nombre_medicion[i].strip():
                        medicion = EquipoMedicion(
                            equipo_codigo=codigo,
                            codigo=codigo_medicion[i],
                            nombre=nombre_medicion[i],
                            ubicacion=ubicacion_medicion[i]
                        )
                        db.session.add(medicion)
                print(f"DEBUG - Antes del commit: {len(db.session.new)} objetos nuevos en sesiÃ³n")
                db.session.commit()
                print(f"DEBUG - Commit ejecutado exitosamente")
                
                # Registrar auditoría
                registrar_auditoria(
                    modulo='equipos',
                    accion='ACTUALIZAR',
                    tabla='equipos',
                    descripcion=f"Editó equipo {equipo.codigo} - {equipo.nombre}"
                )
                
                # Verificar que los motores se guardaron correctamente
                motores_despues = MotorEquipo.query.filter_by(equipo_codigo=codigo).all()
                print(f"DEBUG - DespuÃ©s del commit: {len(motores_despues)} motores en BD")
                for i, motor in enumerate(motores_despues):
                    print(f"DEBUG - Motor en BD {i+1}: id={motor.id}, nomb_Motor='{motor.nomb_Motor}'")
                
                flash(f'Maquina/Equipo "{equipo.codigo}" actualizado correctamente', 'success')
                return redirect(url_for('equipment.listar_equipos'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error al actualizar el equipo: {str(e)}', 'danger')
        else:
            for fieldName, errorMessages in form.errors.items():
                for err in errorMessages:
                    flash(f'Error en {fieldName}: {err}', 'danger')
    return render_template('equipos/editar_equipo.html', form=form, equipo=equipo, motores=motores, mediciones=mediciones)

    #ELIMINAR EQUIPO
@bp.route('/eliminar/<codigo>', methods=['POST'])
@login_required
@require_delete_permission()
def eliminar_equipo(codigo):
    equipo = get_or_404(Equipo, codigo)
    try:
        # Guardar informaciÃ³n antes de eliminar para auditorÃ­a
        equipo_info = f"{equipo.codigo} - {equipo.nombre}"
        
        # Registrar eliminaciÃ³n en el historial
        historial = HistorialEquipo(
            equipo_codigo=codigo,
            tipo_cambio='eliminacion',
            usuario=current_user.username if current_user.is_authenticated else None,
            observaciones=f'Equipo eliminado por el usuario {current_user.username if current_user.is_authenticated else "desconocido"}',
            fecha_cambio=get_local_datetime()
        )
        db.session.add(historial)
        # Eliminar el equipo
        db.session.delete(equipo)
        db.session.commit()
        
        # Registrar auditoría
        registrar_auditoria(
            modulo='equipos',
            accion='ELIMINAR',
            tabla='equipos',
            descripcion=f"Eliminó equipo {equipo_info}"
        )
        
        flash(f'Equipo "{codigo}" eliminado correctamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar el equipo: {str(e)}', 'danger')
    return redirect(url_for('equipment.listar_equipos'))


    #EXPORTAR EL INFORME DE EQUIPOS EN EXCEL
@bp.route('/exportar_equipos', methods=['GET'])
@login_required
@require_any_role('admin', 'supervisor')
def exportar_equipos():
    equipos = Equipo.query.all()

    if not equipos:
        flash("No hay equipos registrados para exportar.", "warning")
        return redirect(url_for('equipment.listar_equipos')) # Redirige a la vista de equipos

    # Crear una lista de diccionarios con los datos de los equipos
    data = [{
        "CÃ³digo": equipo.codigo,
        "Nombre": equipo.nombre,
        "Fecha Ingreso": equipo.fecha_ingreso.strftime('%Y-%m-%d') if equipo.fecha_ingreso else "",
        "Propietario": equipo.propietario,
        "Registro Nuevo": equipo.registro_nuevo,
        "ActualizaciÃ³n": equipo.actualizacion,
        "NÂ° FabricaciÃ³n": equipo.num_fabricacion,
        "Fabricante": equipo.fabricante,
        "Nombre Contacto": equipo.nom_contacto,
        "TelÃ©fono": equipo.telefono,
        "Propietario": equipo.propietario,
        "Tipo Equipo": equipo.tipo_eq,
        "Modelo": equipo.modelo,
        "Tipo de Control": equipo.tipo_control,
        "UbicaciÃ³n": equipo.ubicacion,
        "Clase": equipo.clase,
        "Marca": equipo.marca,
        "Referencia": equipo.referencia,
        "Serie": equipo.serie,
        "Color": equipo.color,
        "Altura": equipo.altura,
        "Largo": equipo.largo,
        "Ancho": equipo.ancho,
        "Peso": equipo.peso,
        "Tipo de EnergÃ­a": equipo.tipo_energia,
        "Corriente": equipo.corriente,
        "Potencia Instalada": equipo.potencia,
        "Voltaje": equipo.voltaje,
        "Tipo Refrigerante": equipo.tipo_refrig ,
        "Tipo Combustible": equipo.tipo_comb ,
        "Tipo Lubricante": equipo.tipo_lub ,
        "Repuestos": equipo.repuestos,
        "Estado Equipo": equipo.estado_eq,
        "NÂ° Motores": equipo.n_motores ,
        "Observaciones": equipo.observaciones,
        "Historial de Mantenimientos": equipo.hist_mtto,
        "FunciÃ³n de la MÃ¡quina": equipo.funcion_maq,
        "OperaciÃ³n": equipo.operacion,
        "MecÃ¡nico": equipo.mecanico,
        "ElÃ©ctrico": equipo.electrico,
        "Partes": equipo.partes,

        # Agrega mÃ¡s campos segÃºn lo que necesites
    } for equipo in equipos]

    # Crear un DataFrame de pandas
    df = pd.DataFrame(data)

    # Crear un buffer para el archivo Excel
    buffer = io.BytesIO()

    # Guardar el DataFrame en el buffer como archivo Excel, pero primero crear encabezado con openpyxl
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Equipos", startrow=7)  # Deja espacio para el encabezado
        workbook = writer.book
        worksheet = writer.sheets["Equipos"]

        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, Border, Side
        from openpyxl.cell.cell import MergedCell
        # --- Encabezado personalizado ---
        # Insertar logo
        logo_path = os.path.join("static", "images", "logo-light.png")
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 110
            img.width = 90
            worksheet.add_image(img, "A1")

        # Bordes
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Combinar celdas para el nombre de la empresa (B1:Q4)
        worksheet.merge_cells('B1:Q4')
        cell_empresa = worksheet['B1']
        cell_empresa.value = "INR INVERSIONES\nREINOSO Y CIA. LTDA."
        cell_empresa.font = Font(bold=True, size=14)
        cell_empresa.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Combinar celdas para el tÃ­tulo (R1:AE4)
        worksheet.merge_cells('R1:AE4')
        cell_titulo = worksheet['R1']
        cell_titulo.value = "HOJA DE VIDA DE EQUIPOS"
        cell_titulo.font = Font(bold=True, size=12)
        cell_titulo.alignment = Alignment(horizontal="center", vertical="center")

        # Tabla vertical a la derecha (AK1:AN1, AK2:AN2, AK3:AN3, AK4:AN4)
        worksheet.merge_cells('AK1:AN1')
        worksheet.merge_cells('AK2:AN2')
        worksheet.merge_cells('AK3:AN3')
        worksheet.merge_cells('AK4:AN4')

        worksheet['AK1'].value = "CÃ³digo"
        worksheet['AK1'].font = Font(bold=True, size=10)
        worksheet['AK1'].alignment = Alignment(horizontal="center", vertical="center")

        worksheet['AK2'].value = "71-MT-56"
        worksheet['AK2'].font = Font(underline="single", color="0000FF")
        worksheet['AK2'].alignment = Alignment(horizontal="center", vertical="center")

        worksheet['AK3'].value = "EdiciÃ³n"
        worksheet['AK3'].font = Font(bold=True, size=10)
        worksheet['AK3'].alignment = Alignment(horizontal="center", vertical="center")

        worksheet['AK4'].value = "25/Jun/2025"
        worksheet['AK4'].font = Font(size=10)
        worksheet['AK4'].alignment = Alignment(horizontal="center", vertical="center")

        # Bordes para la tabla derecha
        for row in worksheet.iter_rows(min_row=1, max_row=4, min_col=37, max_col=40):  # AK=37, AN=40
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.border = border

        # Ajustar ancho de columnas principales del encabezado
        worksheet.column_dimensions['A'].width = 14
        worksheet.column_dimensions['B'].width = 28
        worksheet.column_dimensions['C'].width = 28
        worksheet.column_dimensions['D'].width = 28
        worksheet.column_dimensions['E'].width = 28
        worksheet.column_dimensions['F'].width = 24
        worksheet.column_dimensions['G'].width = 24
        worksheet.column_dimensions['H'].width = 24
        worksheet.column_dimensions['I'].width = 12
        worksheet.column_dimensions['J'].width = 18

    # Regresar el archivo Excel como respuesta de descarga
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="informe_equipos.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    #EXPORTAR INFORME EN PDF
@bp.route('/informe_equipo/<codigo>')
@login_required
def informe_equipo(codigo):
    equipo = get_or_404(Equipo, codigo)
    mantenimientos = Programado.query.filter_by(codigo=codigo).order_by(Programado.fecha_prog.desc()).all()
    historial = HistorialEquipo.query.filter_by(equipo_codigo=codigo).order_by(HistorialEquipo.fecha_cambio.desc()).all()
    # Calcular estadÃ­sticas
    costo_total = sum((mtto.costo_rep or 0) + (mtto.costo_herram or 0) + (mtto.costo_mdo or 0) for mtto in mantenimientos)
    mantenimientos_completados = sum(1 for mtto in mantenimientos if mtto.estado_final == 'Completado')
    # Calcular tiempo promedio
    tiempos_gastados = []
    for mtto in mantenimientos:
        if mtto.hora_inicial and mtto.hora_final:
            tiempo = mtto.hora_final - mtto.hora_inicial
            tiempos_gastados.append(tiempo.total_seconds() / 3600)  # Convertir a horas
    tiempo_promedio = sum(tiempos_gastados) / len(tiempos_gastados) if tiempos_gastados else 0
    # Calcular costos por tipo de mantenimiento
    costos_por_tipo = {}
    for mtto in mantenimientos:
        if mtto.tipo_mantenimiento not in costos_por_tipo:
            costos_por_tipo[mtto.tipo_mantenimiento] = 0
        costos_por_tipo[mtto.tipo_mantenimiento] += (mtto.costo_rep or 0) + (mtto.costo_herram or 0) + (mtto.costo_mdo or 0)
    return render_template('equipos/informe_equipo.html',
                         equipo=equipo,
                         mantenimientos=mantenimientos,
                         costo_total=costo_total,
                         mantenimientos_completados=mantenimientos_completados,
                         tiempo_promedio=tiempo_promedio,
                         costos_por_tipo=costos_por_tipo,
                         historial=historial)

@bp.route('/', methods=['GET'])
def lista_equipos():
    equipos = Equipo.query.all()
    return render_template('equipos/listar_equipos.html', equipos=equipos)

@bp.route('/hoja_vida/<codigo>')
@login_required
def hoja_vida_equipo(codigo):
    equipo = get_or_404(Equipo, codigo)
    # Todos los mantenimientos del equipo
    todos_mantenimientos = Programado.query.filter_by(codigo=codigo).all()
    # Solo completados para la tabla de mantenimientos realizados
    mantenimientos = [m for m in todos_mantenimientos if m.estado_final == 'Completado']
    
    # Calcular totales
    total_mantenimientos = len(todos_mantenimientos)
    total_programados = sum(1 for m in todos_mantenimientos if m.estado_final != 'Completado')
    total_completados = sum(1 for m in todos_mantenimientos if m.estado_final == 'Completado')
    
    # Calcular costo total y tasa de completaciÃ³n
    costo_total = sum((m.costo_rep or 0) + (m.costo_herram or 0) + (m.costo_mdo or 0) for m in mantenimientos)
    total_para_tasa = total_completados + total_programados
    tasa_completacion = (total_completados / total_para_tasa) * 100 if total_para_tasa > 0 else 0
    
    return render_template('equipos/hoja_vida.html', 
                           equipo=equipo, 
                           mantenimientos=mantenimientos, 
                           costo_total=costo_total,
                           tasa_completacion=tasa_completacion,
                           total_mantenimientos=total_mantenimientos,
                           total_programados=total_programados,
                           total_completados=total_completados,
                           es_pdf=False,
                           base_url=request.host_url)

@bp.route('/descargar_hoja_vida/<codigo>')
@login_required
def descargar_hoja_vida(codigo):
    try:
        equipo = get_or_404(Equipo, codigo)
        # Todos los mantenimientos del equipo
        todos_mantenimientos = Programado.query.filter_by(codigo=codigo).all()
        # Solo completados para la tabla de mantenimientos realizados
        mantenimientos = [m for m in todos_mantenimientos if m.estado_final == 'Completado']
        
        # Usar la función que incluye paginación "Página X de Y"
        from utils import generar_y_enviar_pdf_hoja_vida
        return generar_y_enviar_pdf_hoja_vida(equipo, mantenimientos, nombre_archivo=f"hoja_vida_{equipo.codigo}.pdf")
        
    except Exception as e:
        current_app.logger.error(f"Error al generar PDF de hoja de vida para equipo {codigo}: {str(e)}")
        current_app.logger.error(f"Traceback completo: {traceback.format_exc()}")
        flash('Error al generar el PDF de la hoja de vida. Por favor, intente nuevamente.', 'error')
        return redirect(url_for('equipment.hoja_vida_equipo', codigo=codigo))

@bp.route('/importar', methods=['GET', 'POST'])
@login_required
@require_any_role('admin', 'supervisor')
def importar_equipos():
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionÃ³ ningÃºn archivo', 'error')
            return redirect(request.url)
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionÃ³ ningÃºn archivo', 'error')
            return redirect(request.url)
        
        if not archivo.filename.endswith(('.xls', '.xlsx')):
            flash('El archivo debe ser de Excel (.xls o .xlsx)', 'error')
            return redirect(request.url)
        
        try:
            # Leer el archivo Excel
            df = pd.read_excel(archivo)
            
            # Validar columnas requeridas
            columnas_requeridas = ['nombre', 'modelo', 'serie', 'ubicacion']
            if not all(col in df.columns for col in columnas_requeridas):
                flash('El archivo debe contener las columnas: nombre, modelo, serie, ubicacion', 'error')
                return redirect(request.url)
            
            # Procesar cada fila
            equipos_importados = 0
            for _, row in df.iterrows():
                # Verificar si el equipo ya existe por nÃºmero de serie
                if not Equipo.query.filter_by(serie=row['serie']).first():
                    nuevo_equipo = Equipo(
                        nombre=row['nombre'],
                        modelo=row['modelo'],
                        serie=row['serie'],
                        ubicacion=row['ubicacion'],
                        estado='activo',
                        fecha_instalacion=datetime.now()
                    )
                    db.session.add(nuevo_equipo)
                    equipos_importados += 1
            
            db.session.commit()
            flash(f'Se importaron {equipos_importados} equipos exitosamente', 'success')
            return redirect(url_for('equipment.lista_equipos'))
            
        except Exception as e:
            flash(f'Error al importar el archivo: {str(e)}', 'error')
            return redirect(request.url)
    
    return render_template('equipment/importar.html')

@bp.route('/plantilla')
@login_required
@require_any_role('admin', 'supervisor')
def descargar_plantilla():
    # Crear un DataFrame con los campos requeridos y ejemplos
    df = pd.DataFrame({
        'codigo': ['EJEMPLO-001', 'EJEMPLO-002'],
        'nombre': ['Ejemplo Equipo 1', 'Ejemplo Equipo 2'],
        'modelo': ['Modelo A', 'Modelo B'],
        'serie': ['SERIE001', 'SERIE002'],
        'ubicacion': ['UbicaciÃ³n 1', 'UbicaciÃ³n 2']
    })
    
    # Crear un buffer en memoria
    output = io.BytesIO()
    
    # Guardar el DataFrame como Excel en el buffer
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Equipos')
        
        # Obtener el objeto workbook y worksheet
        workbook = writer.book
        worksheet = writer.sheets['Equipos']
        
        # Formato para el encabezado
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D9E1F2',
            'border': 1
        })
        
        # Aplicar formato al encabezado
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            worksheet.set_column(col_num, col_num, 20)
    
    # Mover el puntero al inicio del buffer
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_equipos.xlsx'
    )

@bp.route('/descargar_ficha_tecnica/<codigo>')
@login_required
def descargar_ficha_tecnica(codigo):
    equipo = get_or_404(Equipo, codigo)
    motores = MotorEquipo.query.filter_by(equipo_codigo=codigo).all()
    from utils import generar_y_enviar_pdf_ficha_tecnica
    return generar_y_enviar_pdf_ficha_tecnica(equipo, motores, nombre_archivo=f"ficha_tecnica_{equipo.codigo}.pdf")

@bp.route('/agregar_motor', methods=['POST'])
@login_required
@require_any_role('admin', 'supervisor')
def agregar_motor():
    if not request.form.get('csrf_token'):
        return jsonify({'error': 'Token CSRF invÃ¡lido'}), 400
    
    try:
        # Obtener datos del formulario
        equipo_codigo = request.form.get('equipo_codigo')
        nombre = request.form.get('nombre')
        descripcion = request.form.get('descripcion')
        tipo = request.form.get('tipo')
        rotacion = request.form.get('rotacion')
        rpm = request.form.get('rpm')
        eficiencia = request.form.get('eficiencia')
        corriente = request.form.get('corriente')
        potencia = request.form.get('potencia')
        voltaje = request.form.get('voltaje')
        
        # Crear nuevo motor
        nuevo_motor = MotorEquipo(
            equipo_codigo=equipo_codigo,
            nomb_Motor=nombre,
            descrip_Motor=descripcion,
            tipo_Motor=tipo,
            rotacion=rotacion,
            rpm_Motor=rpm,
            eficiencia=eficiencia,
            corriente_Motor=corriente,
            potencia_Motor=potencia,
            voltaje_Motor=voltaje
        )
        
        db.session.add(nuevo_motor)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Motor agregado exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al agregar motor: {str(e)}'}), 500

@bp.route('/editar_motor/<int:id>', methods=['POST'])
@login_required
@require_any_role('admin', 'supervisor')
def editar_motor(id):
    if not request.form.get('csrf_token'):
        return jsonify({'error': 'Token CSRF invÃ¡lido'}), 400
    
    try:
        motor = MotorEquipo.query.get_or_404(id)
        
        # Actualizar datos del motor
        motor.nomb_Motor = request.form.get('nombre')
        motor.descrip_Motor = request.form.get('descripcion')
        motor.tipo_Motor = request.form.get('tipo')
        motor.rotacion = request.form.get('rotacion')
        motor.rpm_Motor = request.form.get('rpm')
        motor.eficiencia = request.form.get('eficiencia')
        motor.corriente_Motor = request.form.get('corriente')
        motor.potencia_Motor = request.form.get('potencia')
        motor.voltaje_Motor = request.form.get('voltaje')
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Motor actualizado exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al actualizar motor: {str(e)}'}), 500

@bp.route('/eliminar_motor/<int:id>', methods=['POST'])
@login_required
@require_any_role('admin', 'supervisor')
def eliminar_motor(id):
    if not request.form.get('csrf_token'):
        return jsonify({'error': 'Token CSRF invÃ¡lido'}), 400
    
    try:
        motor = MotorEquipo.query.get_or_404(id)
        db.session.delete(motor)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Motor eliminado exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al eliminar motor: {str(e)}'}), 500

@bp.route('/listado_maestro', methods=['GET'])
@login_required
@require_any_role('admin', 'supervisor')
def listado_maestro():
    equipos = Equipo.query.all()
    data = []
    for equipo in equipos:
        # Procesar tipos de energÃ­a
        energias = {'E': 'NO', 'H': 'NO', 'N': 'NO', 'T': 'NO', 'M': 'NO', 'EL': 'NO', 'Q': 'NO'}
        if equipo.tipo_energia:
            for tipo in equipo.tipo_energia.split(','):
                tipo = tipo.strip()
                if tipo == 'ElÃ©ctrica': energias['E'] = 'SI'
                if tipo == 'HidrÃ¡ulica': energias['H'] = 'SI'
                if tipo == 'NeumÃ¡tica': energias['N'] = 'SI'
                if tipo == 'TÃ©rmica': energias['T'] = 'SI'
                if tipo == 'MecÃ¡nica': energias['M'] = 'SI'
                if tipo == 'ElectrÃ³nica': energias['EL'] = 'SI'
                if tipo == 'QuÃ­mica': energias['Q'] = 'SI'
        data.append({
            'Codigo': equipo.codigo,
            'Nombre': equipo.nombre,
            'Tipo equipo': equipo.tipo_eq,
            'Modelo': equipo.modelo,
            'Serie': equipo.serie,
            'UbicaciÃ³n': equipo.ubicacion,
            'Proceso': equipo.proceso,
            'Centro de costo': equipo.centro_costos,
            'Estado': equipo.estado_eq,
            'AÃ±os de operaciÃ³n': equipo.anios_operacion,
            'Fecha de fabricaciÃ³n': equipo.fecha_fabricacion if hasattr(equipo, 'fecha_fabricacion') else '',
            'E: Electrica': energias['E'],
            'H: Hidraulica': energias['H'],
            'N: Neumatica': energias['N'],
            'T: Termica': energias['T'],
            'M: Mecanica': energias['M'],
            'EL: Electronica': energias['EL'],
            'Q: Quimica': energias['Q'],
            'Tipo de inst. electrica': equipo.tipo_instalacion,
            'Funcion de la maquina': equipo.funcion_maq,
            'Operacion': 'SI' if equipo.operacion_file else 'NO',
            'Instructivo': 'SI' if equipo.instructivos_file else 'NO',
            'Estandar de seguridad': 'SI' if equipo.estandar_seguridad_file else 'NO',
            'Cartas de lubricacion': equipo.cartas_lubricacion,
            'Ficha tecnica': 'SI' if equipo.ficha_tecnica else 'NO',
            'Hoja de vida': 'SI' if equipo.hoja_vida else 'NO',
            'Preoperacional': 'SI' if equipo.preoperacional else 'NO',
            'Plan de mantenimiento': 'SI' if equipo.plan_mantenimiento else 'NO',
            'Manuales': 'SI' if equipo.manuales else 'NO',
            'Metodo de codificacion': equipo.metodo_codificacion,
            'Frecuencia de mantenimiento': equipo.frecuencia_mantenimiento
        })
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Listado Maestro')
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='listado_maestro.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Nueva ruta para ver el historial de un equipo
@bp.route('/historial/<codigo>')
@login_required
@require_any_role('super_admin', 'admin', 'supervisor')
def historial_equipo(codigo):
    equipo = get_or_404(Equipo, codigo)
    historial = HistorialEquipo.query.filter_by(equipo_codigo=codigo).order_by(HistorialEquipo.fecha_cambio.desc()).all()
    return render_template('equipos/historial_equipo.html', equipo=equipo, historial=historial)

@bp.route('/historiales', methods=['GET'])
@login_required
@require_any_role('super_admin', 'admin', 'supervisor')
def historiales_equipos():
    codigo = request.args.get('codigo', '').strip()
    if codigo:
        historiales = HistorialEquipo.query.filter(HistorialEquipo.equipo_codigo.ilike(f"%{codigo}%")).order_by(HistorialEquipo.fecha_cambio.desc()).all()
    else:
        historiales = HistorialEquipo.query.order_by(HistorialEquipo.fecha_cambio.desc()).limit(100).all()
    return render_template('equipos/historiales_equipos.html', historiales=historiales, codigo=codigo)
