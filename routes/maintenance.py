from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, current_app, session, send_file, make_response
from flask_login import login_required, current_user
from datetime import datetime, timedelta, time
from models import db, Programado, Equipo, HistorialCambio, Company, User, get_or_404
from forms import MantenimientoForm, MantenimientoEditarForm, EliminarForm
from utils import require_role, require_any_role, can_edit_mantenimiento, can_view_mantenimiento, get_mantenimientos_filtrados_por_rol, get_usuarios_filtrados_por_rol, get_equipos_filtrados_por_rol, require_delete_permission, registrar_auditoria
import pdfkit
import os
import tempfile
from dateutil.relativedelta import relativedelta
from collections import Counter
import pandas as pd
import holidays
from sqlalchemy import or_, and_, func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from functools import wraps
import calendar
from utils import get_pdf_config, get_pdf_options


config = get_pdf_config()

maintenance = Blueprint('maintenance', __name__)

def get_user_display_name(username_or_name):
    if not username_or_name:
        return username_or_name
    
    # Primero intentar buscar como usuario
    user = User.query.filter_by(username=username_or_name, is_active=True).first()
    if user:
        return user.name if user.name else user.username
    
    # Si no es un usuario, verificar si es una empresa
    company = Company.query.filter_by(nombre=username_or_name, activo=True).first()
    if company:
        return company.nombre
    
    # Si no es ni usuario ni empresa, devolver el valor original (puede ser "Otro")
    return username_or_name

@maintenance.route('/equipo/<codigo>', methods=['GET'])
def get_equipo_info(codigo):
    try:
        # Verificar si el usuario está autenticado
        if not current_user.is_authenticated:
            return jsonify({
                'success': False,
                'error': 'Usuario no autenticado'
            }), 401

        current_app.logger.info(f"Buscando equipo con código: {codigo}")
        equipo = Equipo.query.filter_by(codigo=codigo).first()
        
        if equipo:
            current_app.logger.info(f"Equipo encontrado: {equipo.nombre}, Ubicación: {equipo.ubicacion}")
            return jsonify({
                'success': True,
                'equipo': {
                    'codigo': equipo.codigo,
                    'nombre': equipo.nombre,
                    'ubicacion': equipo.ubicacion
                }
            })
        else:
            current_app.logger.warning(f"No se encontró el equipo con código: {codigo}")
            return jsonify({
                'success': False,
                'error': 'Equipo no encontrado'
            }), 404
            
    except Exception as e:
        current_app.logger.error(f"Error al buscar equipo: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@maintenance.route('/programar', methods=['GET', 'POST'])
@login_required
def programar():
    current_app.logger.info(f"Método de la petición: {request.method}")
    
    form = MantenimientoForm()
    
    # Si es POST, forzar el valor de form.codigo.data con el recibido en el formulario
    if request.method == 'POST' and 'codigo' in request.form:
        form.codigo.data = request.form['codigo']
    # Obtener todos los equipos disponibles (sin filtrar inicialmente)
    equipos_disponibles = Equipo.query.all()
    codigos = [e.codigo for e in equipos_disponibles]
    # Si el código seleccionado no está en la lista, agregarlo temporalmente
    if form.codigo.data and form.codigo.data not in codigos:
        form.codigo.choices = [('', 'Seleccione un equipo')] + [(form.codigo.data, form.codigo.data)] + [(e.codigo, e.codigo) for e in equipos_disponibles]
    else:
        form.codigo.choices = [('', 'Seleccione un equipo')] + [(e.codigo, e.codigo) for e in equipos_disponibles]
    
    # Obtener lista de usuarios para los selectores
    usuarios = User.query.filter_by(is_active=True).all()
    form.autorizado_por.choices = [(0, 'Seleccione un usuario')] + [(u.id, u.name or u.username) for u in usuarios]
    form.tecnico_asignado.choices = [(0, 'Seleccione un usuario')] + [(u.id, u.name or u.username) for u in usuarios]
    
    # Crear opciones combinadas para técnico realizador
    opciones_tecnico_realizador = [('', 'Seleccione una opción')]

    # Agregar usuarios activos
    for u in usuarios:
        opciones_tecnico_realizador.append((f"USER_{u.id}", u.name or u.username))
        
    # Agregar empresas activas
    empresas = Company.query.filter_by(activo=True).all()
    for e in empresas:
        opciones_tecnico_realizador.append((f"COMPANY_{e.id}", e.nombre))
            
    # Agregar opción "Otro"
    opciones_tecnico_realizador.append(('OTRO', 'Otro'))
    
    # Configurar opciones para el campo company_id
    form.company_id.choices = [(0, 'Seleccione una empresa')] + [(e.id, e.nombre) for e in empresas]
    # El company_id se establecerá en la sección GET

    # Precargar datos si vienen por GET
    if request.method == 'GET':
        codigo = request.args.get('equipo')
        nombre = request.args.get('nombre')
        ubicacion = request.args.get('ubicacion')
        tipo_mantenimiento = request.args.get('tipo_mantenimiento')
        if codigo:
            form.codigo.data = codigo
        if nombre:
            form.nombre.data = nombre
        if ubicacion:
            form.ubicacion.data = ubicacion
        if tipo_mantenimiento:
            form.tipo_mantenimiento.data = tipo_mantenimiento

    if form.validate_on_submit():
        # Obtener el equipo seleccionado
        equipo = Equipo.query.filter_by(codigo=form.codigo.data).first()
        if not equipo:
            flash('Equipo no encontrado.', 'error')
            return render_template('maintenance/programar.html', form=form, equipos=equipos_disponibles, opciones_tecnico_realizador=opciones_tecnico_realizador, datetime=datetime)

        # Obtener el autorizado por
        autorizado_por = None
        if form.autorizado_por.data and form.autorizado_por.data != 0:
            autorizador = db.session.get(User, form.autorizado_por.data)
            if not autorizador:
                current_app.logger.error(f"Autorizador no encontrado: {form.autorizado_por.data}")
                flash('Autorizador no encontrado', 'error')
                return render_template('maintenance/programar.html', form=form, equipos=equipos_disponibles, opciones_tecnico_realizador=opciones_tecnico_realizador, datetime=datetime)
            autorizado_por = autorizador.username
            current_app.logger.info(f"Autorizado por: {autorizado_por}")

        # Obtener el técnico asignado
        tecnico_asignado = None
        if form.tecnico_asignado.data and form.tecnico_asignado.data != 0:
            tecnico = db.session.get(User, form.tecnico_asignado.data)
            if not tecnico:
                current_app.logger.error(f"Técnico asignado no encontrado: {form.tecnico_asignado.data}")
                flash('Técnico asignado no encontrado', 'error')
                return render_template('maintenance/programar.html', form=form, equipos=equipos_disponibles, opciones_tecnico_realizador=opciones_tecnico_realizador, datetime=datetime)
            tecnico_asignado = tecnico.username
            current_app.logger.info(f"Técnico asignado: {tecnico_asignado}")
        
        # Obtener el técnico realizador
        tecnico_realizador = None
        tecnico_realizador_seleccion = request.form.get('tecnico_realizador_selector', '')
        tecnico_realizador_texto = request.form.get('tecnico_realizador_texto', '')
        nuevo_tecnico_realizador = None
        
        if tecnico_realizador_seleccion:
            if isinstance(tecnico_realizador_seleccion, str) and tecnico_realizador_seleccion.startswith('USER_'):
                # Es un usuario
                user_id = tecnico_realizador_seleccion.replace('USER_', '')
                user = db.session.get(User, user_id)
                if user:
                    nuevo_tecnico_realizador = user.name or user.username
            elif isinstance(tecnico_realizador_seleccion, str) and tecnico_realizador_seleccion.startswith('COMPANY_'):
                # Es una empresa
                company_id = tecnico_realizador_seleccion.replace('COMPANY_', '')
                company = db.session.get(Company, company_id)
                if company:
                    nuevo_tecnico_realizador = company.nombre
            elif tecnico_realizador_seleccion == 'OTRO':
                # Es "Otro" - usar el texto ingresado
                nuevo_tecnico_realizador = tecnico_realizador_texto.strip()
        
        # Si no se especificó técnico realizador y el estado es Completado, asignar automáticamente
        if not nuevo_tecnico_realizador and form.estado_inicial.data == 'Completado':
            if tecnico_asignado:
                # Buscar el usuario asignado para obtener su nombre
                tecnico_asignado_user = User.query.filter_by(username=tecnico_asignado).first()
                if tecnico_asignado_user:
                    nuevo_tecnico_realizador = tecnico_asignado_user.name or tecnico_asignado_user.username
                else:
                    nuevo_tecnico_realizador = tecnico_asignado
                # Solo mostrar el mensaje si el campo estaba vacío antes
                if not tecnico_realizador:
                    flash('El técnico realizador se ha asignado automáticamente con el técnico asignado.', 'info')

        hora_str = form.hora.data.strftime('%H:%M') if form.hora.data else None
        
        # Obtener el tiempo gastado
        if form.hora_inicial.data and form.hora_final.data:
            tdelta = form.hora_final.data - form.hora_inicial.data
            total_minutos = int(tdelta.total_seconds() // 60)
            dias = total_minutos // (24 * 60)
            horas = (total_minutos % (24 * 60)) // 60
            minutos = total_minutos % 60
            
            partes = []
            if dias > 0:
                partes.append(f"{dias} día{'s' if dias > 1 else ''}")
            if horas > 0:
                partes.append(f"{horas}h")
            if minutos > 0 or not partes:
                partes.append(f"{minutos}m")
            
            tiempo_gastado = ' '.join(partes) if partes else "0m"
        else:
            tiempo_gastado = None
        
        # Crear nuevo mantenimiento
        nuevo_mtto = Programado(
            codigo=equipo.codigo,
            nombre=form.nombre.data,
            fecha_prog=form.fecha_prog.data,
            hora=hora_str,
            servicio=form.servicio.data,
            tipo_mantenimiento=form.tipo_mantenimiento.data,
            tiempo=form.tiempo.data,
            repuestos=form.repuestos.data,
            herramientas=form.herramientas.data,
            ubicacion=form.ubicacion.data,
            autorizado_por=autorizado_por,
            estado_inicial='Programado',
            costo_rep=form.costo_rep.data,
            costo_herram=form.costo_herram.data,
            costo_mdo=form.costo_mdo.data,
            frecuencia=form.frecuencia.data,
            prox_mtto=None,
            observaciones=form.observaciones.data,
            tecnico_asignado=tecnico_asignado,
            tecnico_realizador=nuevo_tecnico_realizador,
            hora_inicial=form.hora_inicial.data,
            hora_final=form.hora_final.data,
            tiempo_gastado=tiempo_gastado,
            company_id=form.company_id.data
        )
        nuevo_mtto.calcular_costo_total()
        nuevo_mtto.calcular_prox_mtto()
        
        db.session.add(nuevo_mtto)
        db.session.commit()
        
        # Registrar auditoría
        registrar_auditoria(
            modulo="Equipos",
            accion='CREAR',
            tabla='programado',
            descripcion=f"Creó mantenimiento #{nuevo_mtto.id} para equipo {nuevo_mtto.codigo} - {nuevo_mtto.nombre}"
        )
        
        # Crear mantenimientos futuros basados en la frecuencia (solo para mantenimientos preventivos)
        mantenimientos_futuros_creados = 0
        if (form.frecuencia.data and form.frecuencia.data != 'Seleccionar' and 
            form.tipo_mantenimiento.data == 'Preventivo'):
            mantenimientos_futuros_creados = crear_mantenimientos_futuros(nuevo_mtto)
            if mantenimientos_futuros_creados > 0:
                flash(f'Mantenimiento #{nuevo_mtto.id} para el equipo {nuevo_mtto.codigo} programado correctamente. Se crearon {mantenimientos_futuros_creados} mantenimientos preventivos futuros adicionales para el año {nuevo_mtto.fecha_prog.year}.', 'success')
            else:
                flash(f'Mantenimiento #{nuevo_mtto.id} para el equipo {nuevo_mtto.codigo} programado correctamente.', 'success')
        else:
            flash(f'Mantenimiento #{nuevo_mtto.id} para el equipo {nuevo_mtto.codigo} programado correctamente.', 'success')
        
        return redirect(url_for('maintenance.lista'))
                
    return render_template('maintenance/programar.html', form=form, equipos=equipos_disponibles, opciones_tecnico_realizador=opciones_tecnico_realizador, datetime=datetime)

def actualizar_mantenimientos_vencidos():
    """
    Función para actualizar automáticamente el estado de mantenimientos vencidos.
    Se ejecuta cada vez que se accede a la lista de mantenimientos.
    """
    try:
        hoy = datetime.now().date()
        
        # Buscar todos los mantenimientos que están vencidos pero no están marcados como vencidos
        mantenimientos_vencidos = Programado.query.filter(
            Programado.fecha_prog < hoy,
            Programado.estado_inicial.in_(['Programado', 'Asignado', 'En espera(Repuestos)', 'En proceso']),
            Programado.estado_final.is_(None)  # Solo los que no están completados o cancelados
        ).all()
        
        mantenimientos_actualizados = 0
        for mtto in mantenimientos_vencidos:
            if mtto.estado_inicial != 'Vencido':
                # Registrar el cambio en el historial
                registrar_cambio(mtto, 'Sistema', 'estado_inicial', mtto.estado_inicial, 'Vencido', 'actualización automática')
                mtto.estado_inicial = 'Vencido'
                mantenimientos_actualizados += 1
        
        if mantenimientos_actualizados > 0:
            db.session.commit()
            print(f"Se actualizaron {mantenimientos_actualizados} mantenimientos a estado 'Vencido'")
            
    except Exception as e:
        print(f"Error al actualizar mantenimientos vencidos: {str(e)}")
        db.session.rollback()

@maintenance.route('/lista')
@login_required
def lista():
    # Programar mantenimientos automáticamente al iniciar el año
    hoy = datetime.now()
    if hoy.month == 1 and hoy.day == 1:
        nuevos = programar_mantenimientos_nuevo_ano()
        if nuevos > 0:
            flash(f'Se programaron automáticamente {nuevos} mantenimientos para el año {hoy.year}.', 'success')
    
    # Actualizar automáticamente mantenimientos vencidos
    actualizar_mantenimientos_vencidos()
    
    year = request.args.get('year', type=int)
    if not year:
        year = datetime.now().year
    
    # Obtener el mes del filtro
    month = request.args.get('month', type=int)
    
    # Generar lista de años disponibles (últimos 5 años y próximos 2)
    current_year = datetime.now().year
    years = list(range(current_year - 5, current_year + 3))
    # Filtros adicionales
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    equipo = request.args.get('equipo')
    estado_inicial = request.args.get('estado_inicial')
    estado_inicial_multiple = request.args.get('estado_inicial_multiple')
    ubicacion = request.args.get('ubicacion')

    query = get_mantenimientos_filtrados_por_rol()

    # Filtro por año (por defecto)
    query = query.filter(
        Programado.fecha_prog >= datetime(year, 1, 1),
        Programado.fecha_prog <= datetime(year, 12, 31)
    )
    
    # Filtro por mes
    if month:
        query = query.filter(Programado.fecha_prog >= datetime(year, month, 1))
        query = query.filter(Programado.fecha_prog <= datetime(year, month + 1, 1) - timedelta(days=1))

    # Filtro por rango de fechas
    if fecha_inicio:
        fecha_inicio_date = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        query = query.filter(Programado.fecha_prog >= fecha_inicio_date)
    if fecha_fin:
        fecha_fin_date = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        query = query.filter(Programado.fecha_prog <= fecha_fin_date)

    # Filtro por equipo
    if equipo:
        query = query.filter(Programado.codigo == equipo)
    # Filtro por estado múltiple
    if estado_inicial_multiple:
        estados_a_filtrar = [e.strip().lower() for e in estado_inicial_multiple.split(',')]
        query = query.filter(func.lower(func.trim(Programado.estado_inicial)).in_(estados_a_filtrar))
    # Filtro por estado_inicial
    elif estado_inicial and estado_inicial.strip():  # Verificar que no esté vacío
        if estado_inicial == 'Vencido':
            query = query.filter(or_(
                and_(Programado.fecha_prog < datetime.now().date(), func.lower(func.trim(Programado.estado_inicial)) == 'programado'),
                func.lower(func.trim(Programado.estado_inicial)) == 'vencido'
            ))
        else:
            query = query.filter(func.lower(func.trim(Programado.estado_inicial)) == estado_inicial.lower())
    
    id_filtro = request.args.get('id')
    # Filtro por ID
    if id_filtro:
        try:
            query = query.filter(Programado.id == int(id_filtro))
        except ValueError:
            pass  # Si el valor no es un entero, ignorar el filtro

    # Nuevos filtros
    busqueda = request.args.get('busqueda', '')
    if busqueda:
        query = query.filter(or_(
            Programado.nombre.ilike(f'%{busqueda}%'),
            Programado.servicio.ilike(f'%{busqueda}%'),
            Programado.tecnico_asignado.ilike(f'%{busqueda}%'),
            Programado.tecnico_realizador.ilike(f'%{busqueda}%')
        ))
    tipo_mantenimiento = request.args.get('tipo_mantenimiento')
    frecuencia = request.args.get('frecuencia')
    tecnico = request.args.get('tecnico')
    tecnico_asignado = request.args.get('tecnico_asignado')
    tecnico_realizador = request.args.get('tecnico_realizador')
    costo_min = request.args.get('costo_min', type=float)
    costo_max = request.args.get('costo_max', type=float)
    ordenar_por = request.args.get('ordenar_por', 'id')

    # Filtro por tipo de mantenimiento
    if tipo_mantenimiento:
        query = query.filter(Programado.tipo_mantenimiento == tipo_mantenimiento)

    if ordenar_por == 'id':
        query = query.order_by(Programado.id.desc())
    elif ordenar_por == 'codigo':
        query = query.order_by(Programado.codigo.asc())
    else:
        query = query.order_by(Programado.fecha_prog.desc())

    mantenimientos = query.all()

    # Definir el total de mantenimientos filtrados antes de cualquier uso
    total_mantenimientos = len(mantenimientos)

    # Calcular fechas de próximo mantenimiento para mantenimientos que no las tengan
    for mtto in mantenimientos:
        if mtto.fecha_prog and mtto.frecuencia and not mtto.prox_mtto:
            mtto.calcular_prox_mtto()
    
    # Guardar los cambios en la base de datos
    db.session.commit()

    # Aplicar conversión de nombres a los mantenimientos
    for mtto in mantenimientos:
        mtto.tecnico_asignado_display = get_user_display_name(mtto.tecnico_asignado)
        mtto.tecnico_realizador_display = get_user_display_name(mtto.tecnico_realizador)
        mtto.autorizado_por_display = get_user_display_name(mtto.autorizado_por)

    # Calcular la fecha de hoy solo una vez
    hoy = datetime.now().date()

    # Calcular lista de IDs de mantenimientos vencidos según la lógica unificada
    ids_vencidos = [m.id for m in mantenimientos if (m.estado_inicial == 'Vencido') or (m.fecha_prog and m.fecha_prog < hoy and m.estado_inicial in ['Programado', 'Asignado'])]

    # Calcular estadísticas adicionales usando la lista filtrada y la lógica unificada
    mantenimientos_vencidos = len(ids_vencidos)

    # Estadísticas por tipo de mantenimiento usando la lista filtrada
    tipos_mantenimiento_stats = {}
    for mtto in mantenimientos:
        tipo = mtto.tipo_mantenimiento
        if tipo not in tipos_mantenimiento_stats:
            tipos_mantenimiento_stats[tipo] = 0
        tipos_mantenimiento_stats[tipo] += 1
    for tipo in tipos_mantenimiento_stats:
        count = tipos_mantenimiento_stats[tipo]
        porcentaje = round((count / total_mantenimientos * 100), 1) if total_mantenimientos > 0 else 0
        tipos_mantenimiento_stats[tipo] = {
            'count': count,
            'porcentaje': porcentaje
        }

    # Verificar equipos sin mantenimientos preventivos programados (solo para admin y supervisor)
    equipos_sin_preventivo = []
    año_actual = datetime.now().year
    
    if current_user.is_admin() or current_user.is_supervisor():
        todos_equipos = Equipo.query.all()
    
    for equipo_item in todos_equipos:
        # Verificar si el equipo tiene mantenimientos preventivos programados para el año actual
        mantenimientos_preventivos = Programado.query.filter(
                Programado.codigo == equipo_item.codigo,
            Programado.tipo_mantenimiento == 'Preventivo',
            Programado.fecha_prog >= datetime(año_actual, 1, 1),
            Programado.fecha_prog <= datetime(año_actual, 12, 31)
        ).first()
        
        if not mantenimientos_preventivos:
            # Obtener el último mantenimiento preventivo (si existe)
            ultimo_preventivo = Programado.query.filter(
                    Programado.codigo == equipo_item.codigo,
                Programado.tipo_mantenimiento == 'Preventivo'
            ).order_by(Programado.fecha_prog.desc()).first()
            
            equipos_sin_preventivo.append({
                    'equipo': equipo_item,
                'ultimo_preventivo': ultimo_preventivo,
                'dias_sin_preventivo': (datetime.now().date() - ultimo_preventivo.fecha_prog).days if ultimo_preventivo else None
            })
    
    # Ordenar por días sin mantenimiento preventivo (descendente)
    equipos_sin_preventivo.sort(key=lambda x: x['dias_sin_preventivo'] or 9999, reverse=True)
    
    # Generar alerta si hay equipos sin mantenimientos preventivos
    if equipos_sin_preventivo:
        equipos_lista = ', '.join([f"{eq['equipo'].codigo} ({eq['equipo'].nombre})" for eq in equipos_sin_preventivo[:5]])  # Mostrar solo los primeros 5
        if len(equipos_sin_preventivo) > 5:
            equipos_lista += f" y {len(equipos_sin_preventivo) - 5} más"
        flash(f'⚠️ ALERTA: {len(equipos_sin_preventivo)} equipos sin mantenimientos preventivos programados para {año_actual}: {equipos_lista}', 'warning')

    # Estadísticas para técnicos
    stats_tecnico={
        'programados': len([m for m in mantenimientos if m.tecnico_asignado == current_user.username and m.estado_inicial in ['Programado', 'Asignado', 'En proceso', 'En espera(Repuestos)']]),
        'completados': len([m for m in mantenimientos if m.tecnico_realizador == (current_user.name or current_user.username) and m.estado_final == 'Completado']),
        'total_asignados': len([m for m in mantenimientos if m.tecnico_asignado == current_user.username])
    } if current_user.role == 'tecnico' else None
    
    # Calcular porcentaje de completado
    if current_user.role == 'tecnico' and stats_tecnico:
        total_asignados = stats_tecnico['total_asignados']
        completados = stats_tecnico['completados']
        stats_tecnico['porcentaje_completado'] = round((completados / total_asignados * 100) if total_asignados > 0 else 0, 1)

    # Obtener lista de equipos para los filtros
    equipos_query = get_equipos_filtrados_por_rol()
    equipos = [(e.codigo, e.codigo) for e in equipos_query.order_by(Equipo.codigo).all()]

    # Definir lista de estados para el filtro
    estados = ['Programado', 'Asignado', 'En espera(Repuestos)', 'En proceso', 'Aplazado', 'Pausado', 'Vencido', 'Completado', 'Cancelado']

    # Definir ubicaciones para el filtro
    ubicaciones = [u[0] for u in equipos_query.with_entities(Equipo.ubicacion).distinct().order_by(Equipo.ubicacion).all() if u[0]]

    # Definir tipos de mantenimiento para el filtro
    tipos_mantenimiento = [t[0] for t in query.with_entities(Programado.tipo_mantenimiento).distinct().order_by(Programado.tipo_mantenimiento).all() if t[0]]

    # Definir frecuencias para el filtro
    frecuencias = ['Seleccionar', 'Diario', 'Semanal', 'Quincenal', 'Mensual', 'Bimestral', 'Trimestral', 'Semestral', 'Anual']

    # Lista de técnicos para filtros
    tecnicos = []
    if current_user.is_admin() or current_user.is_supervisor():
        tecnicos = [u.name or u.username for u in User.query.filter_by(role='tecnico', is_active=True).order_by(User.name or User.username).all()]
    else:
        tecnicos = [current_user.name or current_user.username]

    # Calcular estadísticas por estado usando la lista ya filtrada
    estadisticas_estados = {}
    estados_posibles = ['Programado', 'Asignado', 'En espera(Repuestos)', 'En proceso', 'Aplazado', 'Pausado', 'Vencido', 'Completado', 'Cancelado']
    for estado in estados_posibles:
        count = 0
        for mtto in mantenimientos:
            if estado in ['Completado', 'Cancelado']:
                if mtto.estado_final == estado:
                    count += 1
            else:
                if mtto.estado_inicial == estado:
                    count += 1
        if count > 0:
            porcentaje = round((count / total_mantenimientos * 100), 1) if total_mantenimientos > 0 else 0
            estadisticas_estados[estado] = {
                'count': count,
                'porcentaje': porcentaje
            }

    # Calcular mantenimientos de esta semana y este mes usando la lista filtrada
    mantenimientos_esta_semana = len([
        m for m in mantenimientos
        if m.fecha_prog and m.fecha_prog <= hoy + timedelta(days=7) and m.estado_inicial in ['Programado', 'Asignado']
    ])
    mantenimientos_este_mes = len([
        m for m in mantenimientos
        if m.fecha_prog and m.fecha_prog.month == hoy.month and m.estado_inicial in ['Programado', 'Asignado']
    ])

    # ... después de obtener la lista de mantenimientos ...
    dias_alerta = 15  # igual que en home.py
    ids_proximos = [
        m.id for m in mantenimientos
        if m.fecha_prog and hoy <= m.fecha_prog <= hoy + timedelta(days=dias_alerta)
        and m.estado_inicial in ['Programado', 'Asignado']
        and m.id not in ids_vencidos
    ]

    form_eliminar = EliminarForm()

    return render_template(
        'maintenance/lista.html',
        mantenimientos=mantenimientos,
        ids_vencidos=ids_vencidos,
        equipos_sin_preventivo=equipos_sin_preventivo,
        stats_tecnico=stats_tecnico,
        tipos_mantenimiento_stats=tipos_mantenimiento_stats,
        total_mantenimientos=total_mantenimientos,
        year=year,
        years=years,
        month=month,
        pagination=None,  # Si usas paginación, pásala aquí
        form_eliminar=form_eliminar,
        filtro_equipo=equipo, 
        filtro_estado=estado_inicial, 
        filtro_ubicacion=ubicacion,
        filtro_fecha_inicio=fecha_inicio, 
        filtro_fecha_fin=fecha_fin,
        filtro_id=id_filtro,
        filtro_ordenar_por=ordenar_por,
        filtro_busqueda=busqueda,
        filtro_tipo_mantenimiento=tipo_mantenimiento,
        filtro_frecuencia=frecuencia,
        filtro_tecnico=tecnico,
        filtro_tecnico_asignado=tecnico_asignado,
        filtro_tecnico_realizador=tecnico_realizador,
        filtro_costo_min=costo_min,
        filtro_costo_max=costo_max,
        estadisticas_estados=estadisticas_estados,
        mantenimientos_esta_semana=mantenimientos_esta_semana,
        mantenimientos_este_mes=mantenimientos_este_mes,
        ids_proximos=ids_proximos
    )

def registrar_cambio(mantenimiento, usuario, campo, valor_anterior, valor_nuevo, accion):
    if str(valor_anterior) != str(valor_nuevo):
        cambio = HistorialCambio(
            mantenimiento_id=mantenimiento.id,
            usuario=usuario,
            campo=campo,
            valor_anterior=str(valor_anterior),
            valor_nuevo=str(valor_nuevo),
            accion=accion
        )
        db.session.add(cambio)

@maintenance.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@require_any_role('super_admin', 'admin', 'supervisor', 'tecnico')
def editar_mantenimiento(id):
    tecnico_realizador_seleccionado = None  # Inicialización para evitar UnboundLocalError
    mantenimiento = get_or_404(Programado, id)

    # Prevenir que los técnicos editen mantenimientos completados
    if current_user.role == 'tecnico' and mantenimiento.estado_final == 'Completado':
        flash('Un mantenimiento completado no puede ser editado por un técnico.', 'warning')
        return redirect(url_for('maintenance.lista'))

    if not can_edit_mantenimiento(current_user, mantenimiento):
        flash('No tienes permiso para editar este mantenimiento.', 'danger')
        return redirect(url_for('maintenance.lista'))

    form = MantenimientoEditarForm(obj=mantenimiento)

    equipos = get_equipos_filtrados_por_rol()
    form.codigo.choices = [('', 'Seleccione un equipo')] + [(e.codigo, e.codigo) for e in equipos]

    usuarios = User.query.filter_by(is_active=True).all()
    form.autorizado_por.choices = [(0, 'Seleccione un usuario')] + [(u.id, u.name or u.username) for u in usuarios]
    form.tecnico_asignado.choices = [(0, 'Seleccione un usuario')] + [(u.id, u.name or u.username) for u in usuarios]
    
    # Crear opciones combinadas para técnico realizador
    opciones_tecnico_realizador = [('', 'Seleccione una opción')]
    
    # Agregar usuarios activos
    for u in usuarios:
        opciones_tecnico_realizador.append((f"USER_{u.id}", u.name or u.username))
    
    # Agregar empresas activas
    empresas = Company.query.filter_by(activo=True).all()
    for e in empresas:
        opciones_tecnico_realizador.append((f"COMPANY_{e.id}", e.nombre))
    
    # Agregar opción "Otro"
    opciones_tecnico_realizador.append(('OTRO', 'Otro'))
    
    # Configurar opciones para el campo company_id
    form.company_id.choices = [(0, 'Seleccione una empresa')] + [(e.id, e.nombre) for e in empresas]
    form.company_id.data = mantenimiento.company_id if mantenimiento.company_id else 0
    
    if form.validate_on_submit():
        # Validación especial para estado "Asignado"
        if form.estado_inicial.data == 'Asignado':
            tecnico_id = form.tecnico_asignado.data
            company_id = form.company_id.data
            if (not tecnico_id or tecnico_id == 0) and (not company_id or company_id == 0):
                flash('Para asignar un mantenimiento, debe seleccionar un técnico asignado o una empresa.', 'danger')
                return render_template('maintenance/editar.html', form=form, mantenimiento=mantenimiento, datetime=datetime, user_role=current_user.role, opciones_tecnico_realizador=opciones_tecnico_realizador, tecnico_realizador_seleccionado=tecnico_realizador_seleccionado)

        # Capturar el estado ANTES de cualquier modificación
        estado_anterior = mantenimiento.estado_inicial

        campos_modificados = []
        
        # Diccionario para mapear campos del formulario a atributos del modelo
        campos_a_verificar = {
                'codigo': 'codigo', 'nombre': 'nombre', 'fecha_prog': 'fecha_prog', 
                'hora': 'hora', 'servicio': 'servicio', 'tipo_mantenimiento': 'tipo_mantenimiento',
                'tiempo': 'tiempo', 'repuestos': 'repuestos', 'herramientas': 'herramientas',
                'ubicacion': 'ubicacion', 'estado_inicial': 'estado_inicial', 'motivo': 'motivo',
                'costo_rep': 'costo_rep', 'costo_herram': 'costo_herram', 'costo_mdo': 'costo_mdo',
                'frecuencia': 'frecuencia', 'prox_mtto': 'prox_mtto', 'observaciones': 'observaciones',
                'tecnico_realizador': 'tecnico_realizador', 'hora_inicial': 'hora_inicial',
                'hora_final': 'hora_final', 'company_id': 'company_id'
            }

        for campo_form, campo_modelo in campos_a_verificar.items():
                valor_form = getattr(form, campo_form).data
                valor_db = getattr(mantenimiento, campo_modelo)

                # Manejo especial para company_id
                if campo_form == 'company_id':
                    nuevo_valor = valor_form if valor_form and valor_form != 0 else None
                    if mantenimiento.company_id != nuevo_valor:
                        registrar_cambio(mantenimiento, current_user.username, campo_modelo, mantenimiento.company_id, nuevo_valor, 'edición')
                        mantenimiento.company_id = nuevo_valor
                        continue

                # Manejo especial para el campo hora para convertirlo a string
                if campo_form == 'hora' and isinstance(valor_form, time):
                    valor_form = valor_form.strftime('%H:%M')

                # Normalizar tipos para comparación
                if isinstance(valor_db, datetime) and isinstance(valor_form, datetime):
                    valor_db_str = valor_db.strftime('%Y-%m-%d %H:%M:%S') if valor_db else None
                    valor_form_str = valor_form.strftime('%Y-%m-%d %H:%M:%S') if valor_form else None
                    if valor_db_str != valor_form_str:
                        registrar_cambio(mantenimiento, current_user.username, campo_modelo, valor_db, valor_form, 'edición')
                        setattr(mantenimiento, campo_modelo, valor_form)
                elif str(valor_db or '') != str(valor_form or ''):
                    registrar_cambio(mantenimiento, current_user.username, campo_modelo, valor_db, valor_form, 'edición')
                    setattr(mantenimiento, campo_modelo, valor_form)

            # Manejo especial para 'autorizado_por'
                autorizador_id = form.autorizado_por.data
                nuevo_autorizador = None
                if autorizador_id and autorizador_id != 0:
                    autorizador = db.session.get(User, autorizador_id)
                    if autorizador:
                        nuevo_autorizador = autorizador.username
                
                if mantenimiento.autorizado_por != nuevo_autorizador:
                    registrar_cambio(mantenimiento, current_user.username, 'autorizado_por', mantenimiento.autorizado_por, nuevo_autorizador, 'edición')
                mantenimiento.autorizado_por = nuevo_autorizador
                
            # Manejo especial para 'tecnico_asignado'
                tecnico_id = form.tecnico_asignado.data
                nuevo_tecnico = None
                if tecnico_id and tecnico_id != 0:
                    tecnico = db.session.get(User, tecnico_id)
                    if tecnico:
                        nuevo_tecnico = tecnico.username

                if mantenimiento.tecnico_asignado != nuevo_tecnico:
                    registrar_cambio(mantenimiento, current_user.username, 'tecnico_asignado', mantenimiento.tecnico_asignado, nuevo_tecnico, 'edición')
                mantenimiento.tecnico_asignado = nuevo_tecnico

            # Manejo especial para 'tecnico_realizador'
                tecnico_realizador_seleccion = form.tecnico_realizador.data
                tecnico_realizador_texto = request.form.get('tecnico_realizador_texto', '')
                nuevo_tecnico_realizador = None
            
                if tecnico_realizador_seleccion:
                    if isinstance(tecnico_realizador_seleccion, str) and tecnico_realizador_seleccion.startswith('USER_'):
                        # Es un usuario
                        user_id = tecnico_realizador_seleccion.replace('USER_', '')
                    user = db.session.get(User, user_id)
                    if user:
                            nuevo_tecnico_realizador = user.name or user.username
                    elif isinstance(tecnico_realizador_seleccion, str) and tecnico_realizador_seleccion.startswith('COMPANY_'):
                        # Es una empresa
                        company_id = tecnico_realizador_seleccion.replace('COMPANY_', '')
                    company = db.session.get(Company, company_id)
                    if company:
                            nuevo_tecnico_realizador = company.nombre
                    elif tecnico_realizador_seleccion == 'OTRO':
                        # Es "Otro" - usar el texto ingresado
                        nuevo_tecnico_realizador = tecnico_realizador_texto.strip()

            # Si no se especificó técnico realizador y el estado es Completado, asignar automáticamente
                if not nuevo_tecnico_realizador and form.estado_inicial.data == 'Completado':
                    if mantenimiento.tecnico_asignado:
                        # Buscar el usuario asignado para obtener su nombre
                        tecnico_asignado_user = User.query.filter_by(username=mantenimiento.tecnico_asignado).first()
                        if tecnico_asignado_user:
                            nuevo_tecnico_realizador = tecnico_asignado_user.name or tecnico_asignado_user.username
                        else:
                            nuevo_tecnico_realizador = mantenimiento.tecnico_asignado
                        # Solo mostrar el mensaje si el campo estaba vacío antes
                        if not mantenimiento.tecnico_realizador:
                            flash('El técnico realizador se ha asignado automáticamente con el técnico asignado.', 'info')

                if mantenimiento.tecnico_realizador != nuevo_tecnico_realizador:
                    registrar_cambio(mantenimiento, current_user.username, 'tecnico_realizador', mantenimiento.tecnico_realizador, nuevo_tecnico_realizador, 'edición')
                mantenimiento.tecnico_realizador = nuevo_tecnico_realizador

            # Asignar horas y calcular tiempo gastado
                mantenimiento.hora_inicial = form.hora_inicial.data if form.hora_inicial.data else None
                mantenimiento.hora_final = form.hora_final.data if form.hora_final.data else None

                if mantenimiento.hora_inicial and mantenimiento.hora_final:
                    tdelta = mantenimiento.hora_final - mantenimiento.hora_inicial
                    total_minutos = int(tdelta.total_seconds() // 60)
                    dias = total_minutos // (24 * 60)
                    horas = (total_minutos % (24 * 60)) // 60
                    minutos = total_minutos % 60
                
                    partes = []
                    if dias > 0:
                        partes.append(f"{dias} día{'s' if dias > 1 else ''}")
                    if horas > 0:
                        partes.append(f"{horas}h")
                    if minutos > 0 or not partes:
                        partes.append(f"{minutos}m")
                
                    mantenimiento.tiempo_gastado = ' '.join(partes) if partes else "0m"
                else:
                    mantenimiento.tiempo_gastado = None
                
            # Manejo especial para el cambio de estado
        nuevo_estado = form.estado_inicial.data
        if nuevo_estado == 'Completado':
                if not form.hora_inicial.data or not form.hora_final.data:
                    flash('Para completar un mantenimiento, debe registrar la Hora Inicial y la Hora Final.', 'danger')
                    return render_template('maintenance/editar.html', form=form, mantenimiento=mantenimiento, datetime=datetime, user_role=current_user.role, opciones_tecnico_realizador=opciones_tecnico_realizador, tecnico_realizador_seleccionado=tecnico_realizador_seleccionado)
                mantenimiento.estado_final = 'Completado'
        elif nuevo_estado in ['Cancelado', 'Aplazado', 'Pausado']:
                mantenimiento.estado_final = nuevo_estado
        elif nuevo_estado in ['Programado', 'Asignado', 'En proceso', 'En espera(Repuestos)']:
                mantenimiento.estado_final = None

        mantenimiento.calcular_costo_total()
        mantenimiento.calcular_prox_mtto()
            
        # Commit de todos los cambios
        db.session.commit()
        
        # Registrar auditoría
        registrar_auditoria(
            modulo="Equipos",
            accion='ACTUALIZAR',
            tabla='programado',
            registro_id=mantenimiento.id,
            descripcion=f"Editó mantenimiento #{mantenimiento.id} para equipo {mantenimiento.codigo} - {mantenimiento.nombre}"
        )
            
            # Mensaje personalizado para cambio a "Asignado"
        if nuevo_estado == 'Asignado' and estado_anterior != 'Asignado':
                nombre_asignado = "desconocido"
                es_tecnico = False

                # Primero verificar si hay un técnico asignado
                if mantenimiento.tecnico_asignado:
                    user = User.query.filter_by(username=mantenimiento.tecnico_asignado).first()
                    if user:
                        nombre_asignado = user.name or user.username
                        es_tecnico = True
                
                # Si no es un técnico, verificar si hay una empresa asignada
                if not es_tecnico:
                    # Verificar si hay company_id asignado
                    if mantenimiento.company_id:
                        company = db.session.get(Company, mantenimiento.company_id)
                        if company:
                            nombre_asignado = company.nombre
                            es_tecnico = False
                    # Si no hay company_id, verificar el técnico realizador
                    elif mantenimiento.tecnico_realizador:
                        # Buscar si el técnico realizador es una empresa
                        company = Company.query.filter_by(nombre=mantenimiento.tecnico_realizador).first()
                        if company:
                            nombre_asignado = company.nombre
                            es_tecnico = False
                        else:
                            # Si no es una empresa, usar el valor tal como está
                            nombre_asignado = mantenimiento.tecnico_realizador

                if es_tecnico:
                    flash(f'Mantenimiento #{mantenimiento.id} para equipo {mantenimiento.codigo} asignado correctamente al técnico {nombre_asignado}.', 'success')
                else:
                    flash(f'Mantenimiento #{mantenimiento.id} para equipo {mantenimiento.codigo} asignado correctamente a la empresa {nombre_asignado}.', 'success')
        else:
            flash(f'Mantenimiento #{mantenimiento.id} para el equipo {mantenimiento.codigo} actualizado correctamente.', 'success')
            
        # Redirección siempre tras guardar
            return redirect(url_for('maintenance.lista'))
                
    elif request.method == 'GET':
        # Precargar datos en el formulario
        form.codigo.data = mantenimiento.codigo
        form.nombre.data = mantenimiento.nombre
        form.fecha_prog.data = mantenimiento.fecha_prog
        if mantenimiento.hora:
            try:
                form.hora.data = datetime.strptime(mantenimiento.hora, '%H:%M').time()
            except (ValueError, TypeError):
                form.hora.data = None
        form.servicio.data = mantenimiento.servicio
        form.tipo_mantenimiento.data = mantenimiento.tipo_mantenimiento
        form.tiempo.data = mantenimiento.tiempo
        form.repuestos.data = mantenimiento.repuestos
        form.herramientas.data = mantenimiento.herramientas
        form.ubicacion.data = mantenimiento.ubicacion
        
        # Pre-seleccionar autorizado_por
        if mantenimiento.autorizado_por:
            autorizador = User.query.filter_by(username=mantenimiento.autorizado_por).first()
            if autorizador:
                form.autorizado_por.data = autorizador.id

        # Pre-seleccionar tecnico_asignado
        if mantenimiento.tecnico_asignado:
            tecnico = User.query.filter_by(username=mantenimiento.tecnico_asignado).first()
            if tecnico:
                form.tecnico_asignado.data = tecnico.id
        
        # Pre-seleccionar tecnico_realizador
        tecnico_realizador_seleccionado = None
        if mantenimiento.tecnico_realizador:
            # Buscar si es un usuario
            user = User.query.filter_by(username=mantenimiento.tecnico_realizador).first()
            if user:
                tecnico_realizador_seleccionado = f"USER_{user.id}"
            else:
                # Buscar si es una empresa
                company = Company.query.filter_by(nombre=mantenimiento.tecnico_realizador).first()
                if company:
                    tecnico_realizador_seleccionado = f"COMPANY_{company.id}"
                else:
                    # Si no es ni usuario ni empresa, es "Otro"
                    tecnico_realizador_seleccionado = 'OTRO'
        
        # Pre-seleccionar company_id
        form.company_id.data = mantenimiento.company_id if mantenimiento.company_id else 0
        
        # Si se accede con el parámetro completar=true, pre-llenar como Completado
        if request.args.get('completar') == 'true' and current_user.is_tecnico():
            form.estado_inicial.data = 'Completado'
            # Establecer hora actual como hora inicial si no existe
            if not mantenimiento.hora_inicial:
                form.hora_inicial.data = datetime.now()
            flash('Complete los campos de hora inicial y final para finalizar el mantenimiento.', 'info')
        else:
            form.estado_inicial.data = mantenimiento.estado_inicial
        
        form.motivo.data = mantenimiento.motivo
        form.costo_rep.data = mantenimiento.costo_rep
        form.costo_herram.data = mantenimiento.costo_herram
        form.costo_mdo.data = mantenimiento.costo_mdo
        form.frecuencia.data = mantenimiento.frecuencia
        form.prox_mtto.data = mantenimiento.prox_mtto
        form.observaciones.data = mantenimiento.observaciones
        form.tecnico_realizador.data = tecnico_realizador_seleccionado
        form.hora_inicial.data = mantenimiento.hora_inicial
        form.hora_final.data = mantenimiento.hora_final
        
        # Convertir usernames a nombres para mostrar en la plantilla
        mantenimiento.tecnico_asignado_display = get_user_display_name(mantenimiento.tecnico_asignado)
        mantenimiento.tecnico_realizador_display = get_user_display_name(mantenimiento.tecnico_realizador)
        mantenimiento.autorizado_por_display = get_user_display_name(mantenimiento.autorizado_por)
    
        return render_template('maintenance/editar.html', form=form, mantenimiento=mantenimiento, datetime=datetime, user_role=current_user.role, opciones_tecnico_realizador=opciones_tecnico_realizador, tecnico_realizador_seleccionado=tecnico_realizador_seleccionado)

@maintenance.route('/registrar-tiempo/<int:id>', methods=['POST'])
@login_required
def registrar_tiempo(id):
    mantenimiento = get_or_404(Programado, id)
    
    try:
        hora_inicial = request.form.get('hora_inicial')
        hora_final = request.form.get('hora_final')
        tecnico_realizador = request.form.get('tecnico_realizador')
        
        mantenimiento.hora_inicial = hora_inicial
        mantenimiento.hora_final = hora_final
        mantenimiento.tecnico_realizador = tecnico_realizador
        
        # Calcular el tiempo gastado
        mantenimiento.calcular_tiempo_gastado()
        
        # Actualizar el estado a Completado
        mantenimiento.estado_inicial = 'Completado'
        
        db.session.commit()
        
        # Registrar auditoría
        registrar_auditoria(
            modulo="Equipos",
            accion='ACTUALIZAR',
            tabla='programado',
            descripcion=f"Registró tiempo y completó mantenimiento #{mantenimiento.id} para equipo {mantenimiento.codigo}"
        )
        
        flash('Tiempo registrado exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al registrar el tiempo: {str(e)}', 'error')
    
    return redirect(url_for('maintenance.lista'))

@maintenance.route('/eliminar/<int:id>', methods=['POST'])
@login_required
@require_delete_permission()
def eliminar_mantenimiento(id):
    mantenimiento = get_or_404(Programado, id)
    try:
        # Guardar información antes de eliminar para auditoría
        mantenimiento_info = f"#{mantenimiento.id} - {mantenimiento.codigo} - {mantenimiento.nombre}"
        
        db.session.delete(mantenimiento)
        db.session.commit()
        
        # Registrar auditoría
        registrar_auditoria(
            modulo="Equipos",
            accion='ELIMINAR',
            tabla='programado',
            descripcion=f"Eliminó mantenimiento {mantenimiento_info}"
        )
        
        return jsonify({'success': True, 'message': 'Mantenimiento eliminado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@maintenance.route('/imprimir/<int:id>')
@login_required
def imprimir_mantenimiento(id):
    mantenimiento = get_or_404(Programado, id)
    
    # Verificar permisos para ver este mantenimiento
    if not can_view_mantenimiento(current_user, mantenimiento):
        flash('No tienes permisos para ver este mantenimiento.', 'error')
        return redirect(url_for('maintenance.lista'))
    
    return render_template('maintenance/imprimir.html', mantenimientos=[mantenimiento], now=datetime.now())

@maintenance.route('/ver/<int:id>')
@login_required
def ver_mantenimiento(id):
    mantenimiento = get_or_404(Programado, id)
    
    # Verificar permisos para ver este mantenimiento
    if not can_view_mantenimiento(current_user, mantenimiento):
        flash('No tienes permisos para ver este mantenimiento.', 'error')
        return redirect(url_for('maintenance.lista'))
    
    # Agregar información de display para la plantilla
    mantenimiento.tecnico_asignado_display = get_user_display_name(mantenimiento.tecnico_asignado)
    mantenimiento.autorizado_por_display = get_user_display_name(mantenimiento.autorizado_por)
    
    return render_template('maintenance/ver.html', mantenimiento=mantenimiento)

@maintenance.route('/descargar/<int:id>')
@login_required
def descargar_mantenimiento(id):
    try:
        mantenimiento = get_or_404(Programado, id)
        
        # Verificar permisos para ver este mantenimiento
        if not can_view_mantenimiento(current_user, mantenimiento):
            flash('No tienes permisos para descargar este mantenimiento.', 'error')
            return redirect(url_for('maintenance.lista'))
        
        # Convertir username a nombre para mostrar
        mantenimiento.tecnico_asignado_display = get_user_display_name(mantenimiento.tecnico_asignado)
        mantenimiento.tecnico_realizador_display = get_user_display_name(mantenimiento.tecnico_realizador)
        mantenimiento.autorizado_por_display = get_user_display_name(mantenimiento.autorizado_por)
        
        # Generar PDF con ReportLab
        from utils import create_reportlab_pdf_maintenance_detail
        pdf_buffer = create_reportlab_pdf_maintenance_detail(mantenimiento)
        
        # Enviar el archivo
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f'mantenimiento_{mantenimiento.codigo}_{mantenimiento.fecha_prog.strftime("%Y%m%d")}.pdf',
            mimetype='application/pdf'
        )
            
    except Exception as e:
        current_app.logger.error(f"Error al generar PDF: {str(e)}")
        flash('Error al generar el PDF', 'error')
        return redirect(url_for('maintenance.lista'))

@maintenance.route('/imprimir-todos')
@login_required
def imprimir_todos():
    # Obtener los mismos filtros que en la lista
    year = request.args.get('year', type=int)
    if not year:
        year = datetime.now().year
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    equipo = request.args.get('equipo')
    estado = request.args.get('estado')
    tipo_mantenimiento = request.args.get('tipo_mantenimiento')
    
    # Obtener consulta base filtrada por rol
    query = get_mantenimientos_filtrados_por_rol().filter(
        Programado.fecha_prog >= datetime(year, 1, 1),
        Programado.fecha_prog <= datetime(year, 12, 31)
    )
    if fecha_inicio:
        query = query.filter(Programado.fecha_prog >= datetime.strptime(fecha_inicio, '%Y-%m-%d'))
    if fecha_fin:
        query = query.filter(Programado.fecha_prog <= datetime.strptime(fecha_fin, '%Y-%m-%d'))
    if equipo:
        query = query.filter(Programado.codigo == equipo)
    if estado:
        query = query.filter(func.lower(func.trim(Programado.estado_inicial)) == estado.lower())
    if tipo_mantenimiento:
        query = query.filter(Programado.tipo_mantenimiento == tipo_mantenimiento)
    mantenimientos = query.all()
    
    # Convertir usernames a nombres para mostrar en la plantilla
    for mtto in mantenimientos:
        mtto.tecnico_asignado_display = get_user_display_name(mtto.tecnico_asignado)
        mtto.tecnico_realizador_display = get_user_display_name(mtto.tecnico_realizador)
        mtto.autorizado_por_display = get_user_display_name(mtto.autorizado_por)
    
    # Generar PDF con ReportLab
    from utils import create_reportlab_pdf_maintenance_report
    pdf_buffer = create_reportlab_pdf_maintenance_report(mantenimientos, orientation='landscape')
    
    return send_file(pdf_buffer, as_attachment=True, download_name='mantenimientos_programados.pdf', mimetype='application/pdf')

@maintenance.route('/descargar-todos')
@login_required
def descargar_todos():
    # Igual que imprimir_todos pero genera PDF
    year = request.args.get('year', type=int)
    if not year:
        year = datetime.now().year
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    equipo = request.args.get('equipo')
    estado = request.args.get('estado')
    tipo_mantenimiento = request.args.get('tipo_mantenimiento')
    
    # Obtener consulta base filtrada por rol
    query = get_mantenimientos_filtrados_por_rol().filter(
        Programado.fecha_prog >= datetime(year, 1, 1),
        Programado.fecha_prog <= datetime(year, 12, 31)
    )
    if fecha_inicio:
        query = query.filter(Programado.fecha_prog >= datetime.strptime(fecha_inicio, '%Y-%m-%d'))
    if fecha_fin:
        query = query.filter(Programado.fecha_prog <= datetime.strptime(fecha_fin, '%Y-%m-%d'))
    if equipo:
        query = query.filter(Programado.codigo == equipo)
    if estado:
        query = query.filter(func.lower(func.trim(Programado.estado_inicial)) == estado.lower())
    if tipo_mantenimiento:
        query = query.filter(Programado.tipo_mantenimiento == tipo_mantenimiento)
    mantenimientos = query.all()
    
    # Convertir usernames a nombres para mostrar en la plantilla
    for mtto in mantenimientos:
        mtto.tecnico_asignado_display = get_user_display_name(mtto.tecnico_asignado)
        mtto.tecnico_realizador_display = get_user_display_name(mtto.tecnico_realizador)
        mtto.autorizado_por_display = get_user_display_name(mtto.autorizado_por)

    # Generar PDF con ReportLab
    from utils import create_reportlab_pdf_maintenance_report
    pdf_buffer = create_reportlab_pdf_maintenance_report(mantenimientos, orientation='landscape')
    
    return send_file(pdf_buffer, as_attachment=True, download_name='mantenimientos_programados.pdf', mimetype='application/pdf')

@maintenance.route('/historial/<int:id>/<formato>')
@login_required
def descargar_historial(id, formato):
    mantenimiento = get_or_404(Programado, id)
    
    # Verificar permisos para ver este mantenimiento
    if not can_view_mantenimiento(current_user, mantenimiento):
        flash('No tienes permisos para ver el historial de este mantenimiento.', 'error')
        return redirect(url_for('maintenance.lista'))
    
    historial = HistorialCambio.query.filter_by(mantenimiento_id=id).order_by(HistorialCambio.fecha).all()
    if not historial:
        flash('No hay historial para este mantenimiento', 'warning')
        return redirect(url_for('maintenance.editar_mantenimiento', id=id))
    # Preparar datos
    data = [
        {
            'Fecha': h.fecha.strftime('%Y-%m-%d %H:%M'),
            'Usuario': h.usuario,
            'Campo': h.campo,
            'Valor Anterior': h.valor_anterior,
            'Valor Nuevo': h.valor_nuevo,
            'Acción': h.accion
        }
        for h in historial
    ]
    if formato == 'excel':
        df = pd.DataFrame(data)
        # Agregar código y nombre si hay historial
        if historial and historial[0].mantenimiento:
            codigo = historial[0].mantenimiento.codigo
            nombre = historial[0].mantenimiento.nombre
            encabezado = pd.DataFrame({
                'Fecha': [f'Código: {codigo}'],
                'Usuario': [f'Nombre: {nombre}'],
                'Campo': [''],
                'Valor Anterior': [''],
                'Valor Nuevo': [''],
                'Acción': ['']
            })
            df = pd.concat([encabezado, df], ignore_index=True)
        output = pd.ExcelWriter('historial_temp.xlsx', engine='xlsxwriter')
        df.to_excel(output, index=False, sheet_name='Historial')
        output.close()
        with open('historial_temp.xlsx', 'rb') as f:
            response = make_response(f.read())
        os.remove('historial_temp.xlsx')
        response.headers['Content-Disposition'] = f'attachment; filename=historial_mantenimiento_{id}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
    elif formato == 'pdf':
        # Generar PDF con ReportLab
        from utils import create_reportlab_pdf_historial
        pdf_buffer = create_reportlab_pdf_historial(historial, id)
        response = send_file(pdf_buffer, as_attachment=True, download_name=f'historial_mantenimiento_{id}.pdf', mimetype='application/pdf')
        return response
    else:
        flash('Formato no válido', 'error')
        return redirect(url_for('maintenance.editar_mantenimiento', id=id))

@maintenance.route('/historial/<int:id>/ver')
@login_required
def ver_historial(id):
    mantenimiento = get_or_404(Programado, id)
    
    # Verificar permisos para ver este mantenimiento
    if not can_view_mantenimiento(current_user, mantenimiento):
        flash('No tienes permisos para ver el historial de este mantenimiento.', 'error')
        return redirect(url_for('maintenance.lista'))
    
    historial = HistorialCambio.query.filter_by(mantenimiento_id=id).order_by(HistorialCambio.fecha.desc()).all()
    return render_template('maintenance/historial_ver.html', historial=historial, mantenimiento=mantenimiento)

def programar_mantenimientos_nuevo_ano(año_actual=None):
    from datetime import datetime, timedelta
    import holidays
    from models import Programado, db
    if año_actual is None:
        año_actual = datetime.now().year
    año_anterior = año_actual - 1
    festivos = holidays.country_holidays('CO', years=[año_actual])
    # Buscar mantenimientos programados del año anterior
    mantenimientos_ant = Programado.query.filter(
        Programado.fecha_prog >= datetime(año_anterior, 1, 1),
        Programado.fecha_prog <= datetime(año_anterior, 12, 31),
        Programado.estado_inicial == 'Programado'
    ).all()
    nuevos = 0
    for mtto in mantenimientos_ant:
        # Solo programar mantenimientos preventivos
        if mtto.tipo_mantenimiento != 'Preventivo':
            continue
        # Usar prox_mtto si existe y es del año nuevo
        if mtto.prox_mtto and mtto.prox_mtto.year == año_actual:
            nueva_fecha = mtto.prox_mtto
        else:
            nueva_fecha = mtto.fecha_prog.replace(year=año_actual)
        # Evitar la primera semana de enero (crear después del 6 de enero)
        if nueva_fecha.month == 1 and nueva_fecha.day <= 6:
            nueva_fecha = nueva_fecha.replace(day=7)
        # Evitar domingos y festivos
        while nueva_fecha.weekday() == 6 or nueva_fecha in festivos or (nueva_fecha.month == 1 and nueva_fecha.day <= 6):
            nueva_fecha += timedelta(days=1)
        # Verificar que no exista ya un mantenimiento igual (mismo equipo, fecha, tipo y servicio)
        existe = Programado.query.filter_by(
            codigo=mtto.codigo,
            fecha_prog=nueva_fecha,
            tipo_mantenimiento=mtto.tipo_mantenimiento,
            servicio=mtto.servicio
        ).first()
        if existe:
            continue
        nuevo_mtto = Programado(
            codigo=mtto.codigo,
            nombre=mtto.nombre,
            fecha_prog=nueva_fecha,
            hora=mtto.hora,
            servicio=mtto.servicio,
            tipo_mantenimiento=mtto.tipo_mantenimiento,
            tiempo=mtto.tiempo,
            repuestos=mtto.repuestos,
            herramientas=mtto.herramientas,
            ubicacion=mtto.ubicacion,
            autorizado_por=mtto.autorizado_por,
            estado_inicial='Programado',
            costo_rep=mtto.costo_rep,
            costo_herram=mtto.costo_herram,
            costo_mdo=mtto.costo_mdo,
            frecuencia=mtto.frecuencia,
            prox_mtto=None,
            observaciones=mtto.observaciones,
            tecnico_asignado=mtto.tecnico_asignado,
            tecnico_realizador=mtto.tecnico_realizador,
            hora_inicial=None,
            hora_final=None,
            tiempo_gastado=None,
            company_id=mtto.company_id
        )
        nuevo_mtto.calcular_costo_total()
        db.session.add(nuevo_mtto)
        db.session.commit()
        # Generar futuros según la frecuencia
        from routes.maintenance import crear_mantenimientos_futuros
        nuevos += 1
        nuevos += crear_mantenimientos_futuros(nuevo_mtto)
    if nuevos > 0:
        db.session.commit()
    return nuevos

@maintenance.route('/equipos-sin-preventivo')
@login_required
@require_any_role('super_admin', 'admin', 'supervisor')
def equipos_sin_preventivo():
    # Obtener todos los equipos
    equipos = Equipo.query.all()
    año_actual = datetime.now().year
    equipos_sin_preventivo = []
    
    for equipo in equipos:
        # Verificar si el equipo tiene mantenimientos preventivos programados para el año actual
        mantenimientos_preventivos = Programado.query.filter(
            Programado.codigo == equipo.codigo,
            Programado.tipo_mantenimiento == 'Preventivo',
            Programado.fecha_prog >= datetime(año_actual, 1, 1),
            Programado.fecha_prog <= datetime(año_actual, 12, 31)
        ).first()
        
        if not mantenimientos_preventivos:
            # Obtener el último mantenimiento preventivo (si existe)
            ultimo_preventivo = Programado.query.filter(
                Programado.codigo == equipo.codigo,
                Programado.tipo_mantenimiento == 'Preventivo'
            ).order_by(Programado.fecha_prog.desc()).first()
            
            equipos_sin_preventivo.append({
                'equipo': equipo,
                'ultimo_preventivo': ultimo_preventivo,
                'dias_sin_preventivo': (datetime.now().date() - ultimo_preventivo.fecha_prog).days if ultimo_preventivo else None
            })
    
    # Ordenar por días sin mantenimiento preventivo (descendente)
    equipos_sin_preventivo.sort(key=lambda x: x['dias_sin_preventivo'] or 9999, reverse=True)
    
    total_equipos = Equipo.query.count()
    equipos_con_preventivo = total_equipos - len(equipos_sin_preventivo)
    return render_template('maintenance/equipos_sin_preventivo.html', 
                         equipos_sin_preventivo=equipos_sin_preventivo,
                         año_actual=año_actual,
                         total_equipos=total_equipos,
                         equipos_con_preventivo=equipos_con_preventivo)

@maintenance.route('/equipos-sin-preventivo-json')
@login_required
@require_any_role('super_admin', 'admin', 'supervisor')
def equipos_sin_preventivo_json():
    # Obtener todos los equipos
    equipos = Equipo.query.all()
    año_actual = datetime.now().year
    equipos_sin_preventivo = []
    
    for equipo in equipos:
        # Verificar si el equipo tiene mantenimientos preventivos programados para el año actual
        mantenimientos_preventivos = Programado.query.filter(
            Programado.codigo == equipo.codigo,
            Programado.tipo_mantenimiento == 'Preventivo',
            Programado.fecha_prog >= datetime(año_actual, 1, 1),
            Programado.fecha_prog <= datetime(año_actual, 12, 31)
        ).first()
        
        if not mantenimientos_preventivos:
            # Obtener el último mantenimiento preventivo (si existe)
            ultimo_preventivo = Programado.query.filter(
                Programado.codigo == equipo.codigo,
                Programado.tipo_mantenimiento == 'Preventivo'
            ).order_by(Programado.fecha_prog.desc()).first()
        
            equipos_sin_preventivo.append({
                'codigo': equipo.codigo,
                'nombre': equipo.nombre,
                'ubicacion': equipo.ubicacion,
                'ultimo_preventivo': ultimo_preventivo.fecha_prog.strftime('%d/%m/%Y') if ultimo_preventivo else 'Nunca',
                'dias_sin_preventivo': (datetime.now().date() - ultimo_preventivo.fecha_prog).days if ultimo_preventivo else None
            })
    
    # Ordenar por días sin mantenimiento preventivo (descendente)
    equipos_sin_preventivo.sort(key=lambda x: x['dias_sin_preventivo'] or 9999, reverse=True)
    
    return jsonify(equipos_sin_preventivo)

@maintenance.route('/todos-equipos-json')
@login_required
@require_any_role('super_admin', 'admin', 'supervisor')
def todos_equipos_json():
    equipos = Equipo.query.all()
    equipos_data = []
    for equipo in equipos:
        if equipo.codigo == '0000000000000000000000000000000000000000':
            continue
        equipos_data.append({
            'codigo': equipo.codigo,
            'nombre': equipo.nombre,
            'ubicacion': equipo.ubicacion
        })
    
    return jsonify(equipos_data)

@maintenance.route('/verificar-preventivo/<codigo>')
@login_required
@require_any_role('super_admin', 'admin', 'supervisor')
def verificar_preventivo(codigo):
    try:
        equipo = Equipo.query.filter_by(codigo=codigo).first()
        if not equipo:
            return jsonify({'success': False, 'error': 'Equipo no encontrado'})
        
        # Obtener parámetros opcionales para validación específica
        servicio = request.args.get('servicio')
        frecuencia = request.args.get('frecuencia')
        
        año_actual = datetime.now().year
        query = Programado.query.filter(
            Programado.codigo == codigo,
            Programado.tipo_mantenimiento == 'Preventivo',
            Programado.fecha_prog >= datetime(año_actual, 1, 1),
            Programado.fecha_prog <= datetime(año_actual, 12, 31)
        )
        
        # Si se especifica servicio y frecuencia, buscar mantenimientos que coincidan exactamente
        if servicio and frecuencia:
            mantenimientos_preventivos = query.filter(
                Programado.servicio == servicio,
                Programado.frecuencia == frecuencia
            ).all()
            mensaje = f"Ya existe un mantenimiento preventivo para el equipo {codigo} con el servicio '{servicio}' y frecuencia '{frecuencia}' en el año {año_actual}."
        else:
            # Si no se especifican, mostrar todos los mantenimientos preventivos del equipo
            mantenimientos_preventivos = query.all()
            mensaje = f"El equipo {codigo} tiene {len(mantenimientos_preventivos)} mantenimiento(s) preventivo(s) programado(s) para el año {año_actual}."
        
        return jsonify({
            'success': True,
            'tiene_preventivo': len(mantenimientos_preventivos) > 0,
            'cantidad': len(mantenimientos_preventivos),
            'mensaje': mensaje,
            'mantenimientos': [
                {
                    'id': m.id,
                    'fecha': m.fecha_prog.strftime('%d/%m/%Y'),
                    'estado': m.estado_inicial,
                    'servicio': m.servicio,
                    'frecuencia': m.frecuencia
                } for m in mantenimientos_preventivos
            ]
        })
        
    except Exception as e:
        current_app.logger.error(f"Error al verificar preventivo: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

@maintenance.route('/descargar-informe-excel')
@login_required
def descargar_informe_excel():
    # Obtener los mismos filtros que en la lista
    year = request.args.get('year', type=int)
    if not year:
        year = datetime.now().year
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    equipo = request.args.get('equipo')
    estado = request.args.get('estado')
    tipo_mantenimiento = request.args.get('tipo_mantenimiento')
    
    # Obtener consulta base filtrada por rol
    query = get_mantenimientos_filtrados_por_rol().filter(
        Programado.fecha_prog >= datetime(year, 1, 1),
        Programado.fecha_prog <= datetime(year, 12, 31)
    )
    if fecha_inicio:
        query = query.filter(Programado.fecha_prog >= datetime.strptime(fecha_inicio, '%Y-%m-%d'))
    if fecha_fin:
        query = query.filter(Programado.fecha_prog <= datetime.strptime(fecha_fin, '%Y-%m-%d'))
    if equipo:
        query = query.filter(Programado.codigo == equipo)
    if estado:
        query = query.filter(func.lower(func.trim(Programado.estado_inicial)) == estado.lower())
    if tipo_mantenimiento:
        query = query.filter(Programado.tipo_mantenimiento == tipo_mantenimiento)
    
    mantenimientos = query.all()
    
    # Convertir usernames a nombres para mostrar en la plantilla
    for mtto in mantenimientos:
        mtto.tecnico_asignado_display = get_user_display_name(mtto.tecnico_asignado)
        mtto.tecnico_realizador_display = get_user_display_name(mtto.tecnico_realizador)
        mtto.autorizado_por_display = get_user_display_name(mtto.autorizado_por)

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Mantenimientos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = [
        'ID', 'Código', 'Nombre', 'Fecha Programada', 'Hora', 'Servicio', 
        'Tipo Mantenimiento', 'Estado', 'Técnico Asignado', 'Técnico Realizador',
        'Ubicación', 'Repuestos', 'Herramientas', 'Costo Repuestos', 
        'Costo Herramientas', 'Costo MDO', 'Costo Total', 'Observaciones'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos
    for row, mtto in enumerate(mantenimientos, 2):
        ws.cell(row=row, column=1, value=mtto.id)
        ws.cell(row=row, column=2, value=mtto.codigo)
        ws.cell(row=row, column=3, value=mtto.nombre)
        ws.cell(row=row, column=4, value=mtto.fecha_prog.strftime('%d/%m/%Y') if mtto.fecha_prog else '')
        ws.cell(row=row, column=5, value=mtto.hora)
        ws.cell(row=row, column=6, value=mtto.servicio)
        ws.cell(row=row, column=7, value=mtto.tipo_mantenimiento)
        ws.cell(row=row, column=8, value=mtto.estado_inicial)
        ws.cell(row=row, column=9, value=mtto.tecnico_asignado_display)
        ws.cell(row=row, column=10, value=mtto.tecnico_realizador_display)
        ws.cell(row=row, column=11, value=mtto.ubicacion)
        ws.cell(row=row, column=12, value=mtto.repuestos)
        ws.cell(row=row, column=13, value=mtto.herramientas)
        ws.cell(row=row, column=14, value=mtto.costo_rep or 0)
        ws.cell(row=row, column=15, value=mtto.costo_herram or 0)
        ws.cell(row=row, column=16, value=mtto.costo_mdo or 0)
        ws.cell(row=row, column=17, value=mtto.costo_total or 0)
        ws.cell(row=row, column=18, value=mtto.observaciones)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar el archivo
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'mantenimientos_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@maintenance.route('/estadisticas-tecnicos')
@login_required
@require_any_role('super_admin', 'admin', 'supervisor')
def estadisticas_tecnicos():
    """Vista de estadísticas de completado por técnico para administradores y supervisores"""
    
    year = request.args.get('year', type=int)
    if not year:
        year = datetime.now().year
    
    # Obtener todos los técnicos activos
    tecnicos = User.query.filter_by(role='tecnico', is_active=True).order_by(User.username).all()
    
    # Obtener mantenimientos del año
    mantenimientos = Programado.query.filter(
        Programado.fecha_prog >= datetime(year, 1, 1),
        Programado.fecha_prog <= datetime(year, 12, 31)
    ).all()
    
    # Calcular estadísticas por técnico
    estadisticas = []
    for tecnico in tecnicos:
        # Mantenimientos asignados al técnico
        asignados = [m for m in mantenimientos if m.tecnico_asignado == tecnico.username]
        
        # Mantenimientos completados por el técnico
        completados = [m for m in mantenimientos if m.tecnico_realizador == (tecnico.name or tecnico.username) and m.estado_final == 'Completado']
        
        # Mantenimientos en proceso
        en_proceso = [m for m in mantenimientos if m.tecnico_asignado == tecnico.username and m.estado_inicial in ['En proceso', 'Asignado']]
        
        # Mantenimientos vencidos
        vencidos = [m for m in mantenimientos if m.tecnico_asignado == tecnico.username and m.estado_inicial == 'Vencido']
        
        # Calcular porcentaje de completado
        porcentaje = round((len(completados) / len(asignados) * 100) if len(asignados) > 0 else 0, 1)
        
        # Calcular tiempo promedio de completado
        tiempo_promedio = None
        if completados:
            tiempos = []
            for mtto in completados:
                if mtto.hora_inicial and mtto.hora_final:
                    tiempo = mtto.hora_final - mtto.hora_inicial
                    tiempos.append(tiempo.total_seconds() / 3600)  # en horas
            if tiempos:
                tiempo_promedio = round(sum(tiempos) / len(tiempos), 2)
        
        estadisticas.append({
            'tecnico': tecnico.name or tecnico.username,
            'asignados': len(asignados),
            'completados': len(completados),
            'en_proceso': len(en_proceso),
            'vencidos': len(vencidos),
            'porcentaje': porcentaje,
            'tiempo_promedio': tiempo_promedio,
            'ultimo_completado': max([m.hora_final for m in completados if m.hora_final], default=None)
        })
    
    # Ordenar por porcentaje de completado (descendente)
    estadisticas.sort(key=lambda x: x['porcentaje'], reverse=True)
    
    total_equipos = Equipo.query.count()
    
    # Calcular totales generales
    total_programados = len(mantenimientos)
    total_completados = sum(1 for m in mantenimientos if m.estado_final == 'Completado')
    
    return render_template('maintenance/estadisticas_tecnicos.html', 
                         estadisticas=estadisticas, 
                         year=year,
                         tecnicos=tecnicos,
                         total_equipos=total_equipos,
                         total_programados=total_programados,
                         total_completados=total_completados)

def crear_mantenimientos_futuros(mantenimiento_base):
    """
    Crea múltiples mantenimientos futuros basados en la frecuencia del mantenimiento base.
    Solo programa para el año en curso, evitando domingos y festivos.
    SOLO funciona con mantenimientos de tipo PREVENTIVO.
    Si el siguiente mantenimiento cae en el año siguiente, no lo crea y solo actualiza el campo prox_mtto del último mantenimiento creado.
    
    Args:
        mantenimiento_base: El mantenimiento original que sirve como plantilla
    
    Returns:
        int: Número de mantenimientos creados
    """
    from datetime import datetime, timedelta
    import holidays
    
    # Verificar que sea un mantenimiento preventivo
    if mantenimiento_base.tipo_mantenimiento != 'Preventivo':
        return 0  # No crear futuros para otros tipos de mantenimiento
    
    mantenimientos_creados = 0
    fecha_actual = mantenimiento_base.fecha_prog
    año_actual = fecha_actual.year
    
    # Obtener festivos del año actual
    festivos = holidays.country_holidays('CO', years=[año_actual])
    
    # Calcular el delta de tiempo basado en la frecuencia
    if mantenimiento_base.frecuencia == 'Diario':
        delta = timedelta(days=1)
    elif mantenimiento_base.frecuencia == 'Semanal':
        delta = timedelta(weeks=1)
    elif mantenimiento_base.frecuencia == 'Quincenal':
        delta = timedelta(days=15)
    elif mantenimiento_base.frecuencia == 'Mensual':
        def siguiente_mes(fecha):
            mes = fecha.month % 12 + 1
            anio = fecha.year + (fecha.month // 12)
            dia = min(fecha.day, calendar.monthrange(anio, mes)[1])
            return fecha.replace(year=anio, month=mes, day=dia)
        fecha_siguiente = siguiente_mes(fecha_actual)
    elif mantenimiento_base.frecuencia == 'Bimestral':
        def siguiente_bimestre(fecha):
            mes = (fecha.month - 1 + 2) % 12 + 1
            anio = fecha.year + ((fecha.month - 1 + 2) // 12)
            dia = min(fecha.day, calendar.monthrange(anio, mes)[1])
            return fecha.replace(year=anio, month=mes, day=dia)
        fecha_siguiente = siguiente_bimestre(fecha_actual)
    elif mantenimiento_base.frecuencia == 'Trimestral':
        def siguiente_trimestre(fecha):
            mes = (fecha.month - 1 + 3) % 12 + 1
            anio = fecha.year + ((fecha.month - 1 + 3) // 12)
            dia = min(fecha.day, calendar.monthrange(anio, mes)[1])
            return fecha.replace(year=anio, month=mes, day=dia)
        fecha_siguiente = siguiente_trimestre(fecha_actual)
    elif mantenimiento_base.frecuencia == 'Semestral':
        def siguiente_semestre(fecha):
            mes = (fecha.month - 1 + 6) % 12 + 1
            anio = fecha.year + ((fecha.month - 1 + 6) // 12)
            dia = min(fecha.day, calendar.monthrange(anio, mes)[1])
            return fecha.replace(year=anio, month=mes, day=dia)
        fecha_siguiente = siguiente_semestre(fecha_actual)
    elif mantenimiento_base.frecuencia == 'Anual':
        fecha_siguiente = fecha_actual.replace(year=fecha_actual.year + 1)
    else:
        return 0  # Frecuencia no válida
    
    # Fecha límite: 31 de diciembre del año actual
    fecha_limite = datetime(año_actual, 12, 31).date()
    
    # Para frecuencias que usan días, inicializar fecha_siguiente
    if mantenimiento_base.frecuencia in ['Diario', 'Semanal', 'Quincenal']:
        fecha_siguiente = fecha_actual + delta
    
    ultimo_mtto = mantenimiento_base
    
    while fecha_siguiente <= fecha_limite:
        # Evitar domingos, festivos y la primera semana de enero (después del 6)
        while (fecha_siguiente.weekday() == 6 or 
               fecha_siguiente in festivos or 
               (fecha_siguiente.month == 1 and fecha_siguiente.day <= 6)):
            fecha_siguiente += timedelta(days=1)
        
        # Si la fecha siguiente ya es de otro año, solo actualizar prox_mtto y salir
        if fecha_siguiente.year > año_actual:
            ultimo_mtto.prox_mtto = fecha_siguiente
            db.session.commit()
            break
        
        # Verificar que no exista ya un mantenimiento igual (mismo equipo, fecha, tipo y servicio)
        existe = Programado.query.filter_by(
            codigo=mantenimiento_base.codigo,
            fecha_prog=fecha_siguiente,
            tipo_mantenimiento=mantenimiento_base.tipo_mantenimiento,
            servicio=mantenimiento_base.servicio
        ).first()
        
        if not existe:
            # Crear el nuevo mantenimiento
            nuevo_mtto = Programado(
                codigo=mantenimiento_base.codigo,
                nombre=mantenimiento_base.nombre,
                fecha_prog=fecha_siguiente,
                hora=mantenimiento_base.hora,
                servicio=mantenimiento_base.servicio,
                tipo_mantenimiento=mantenimiento_base.tipo_mantenimiento,
                tiempo=mantenimiento_base.tiempo,
                repuestos=mantenimiento_base.repuestos,
                herramientas=mantenimiento_base.herramientas,
                ubicacion=mantenimiento_base.ubicacion,
                autorizado_por=mantenimiento_base.autorizado_por,
                estado_inicial='Programado',
                costo_rep=mantenimiento_base.costo_rep,
                costo_herram=mantenimiento_base.costo_herram,
                costo_mdo=mantenimiento_base.costo_mdo,
                frecuencia=mantenimiento_base.frecuencia,
                prox_mtto=None,
                observaciones=mantenimiento_base.observaciones,
                tecnico_asignado=mantenimiento_base.tecnico_asignado,
                tecnico_realizador=mantenimiento_base.tecnico_realizador,
                hora_inicial=None,
                hora_final=None,
                tiempo_gastado=None,
                company_id=mantenimiento_base.company_id
            )
            nuevo_mtto.calcular_costo_total()
            nuevo_mtto.calcular_prox_mtto()
            db.session.add(nuevo_mtto)
            mantenimientos_creados += 1
            ultimo_mtto = nuevo_mtto
        else:
            ultimo_mtto = existe
        
        # Calcular la siguiente fecha
        if mantenimiento_base.frecuencia == 'Diario':
            fecha_siguiente += timedelta(days=1)
        elif mantenimiento_base.frecuencia == 'Semanal':
            fecha_siguiente += timedelta(weeks=1)
        elif mantenimiento_base.frecuencia == 'Quincenal':
            fecha_siguiente += timedelta(days=15)
        elif mantenimiento_base.frecuencia == 'Mensual':
            fecha_siguiente = siguiente_mes(fecha_siguiente)
        elif mantenimiento_base.frecuencia == 'Bimestral':
            fecha_siguiente = siguiente_bimestre(fecha_siguiente)
        elif mantenimiento_base.frecuencia == 'Trimestral':
            fecha_siguiente = siguiente_trimestre(fecha_siguiente)
        elif mantenimiento_base.frecuencia == 'Semestral':
            fecha_siguiente = siguiente_semestre(fecha_siguiente)
        elif mantenimiento_base.frecuencia == 'Anual':
            fecha_siguiente = fecha_siguiente.replace(year=fecha_siguiente.year + 1)
    
    # Hacer commit de todos los mantenimientos creados
    if mantenimientos_creados > 0:
        db.session.commit()
    
    return mantenimientos_creados
